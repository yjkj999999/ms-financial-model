#!/usr/bin/env python3
"""
Morgan Stanley Style Financial Model Generator (DCF) - V7 SOTP Enhanced
=================================================================
Generates investment-grade DCF valuation models with:
- Cover sheet with business overview, scenario summary, and valuation summary
- 3 scenario DCF sheets (Bear / Base / Bull) with WACC cross-sheet references
- WACC decomposition with formula-driven Ke and WACC
- Sensitivity analysis (4 matrices) with yellow cross-highlight
- Comparable companies (with multi-market grouping support)
- SOTP (Sum-of-the-Parts) segment valuation
- Key operational KPIs dashboard
- Historical PE Band analysis

V7 Enhancements (inspired by Xiaomi financial model):
- SOTP valuation: per-segment EV with different multiples/methods
- Operational KPIs: delivery volume, user metrics, launch counts
- Historical PE Band: 3-5 year PE trajectory with median annotation
- Comparable companies: multi-market grouping (A/H/US) with per-market medians
- All new features triggered by optional data fields (backward compatible)

Usage:
    python morgan_stanley_xlsx.py output.xlsx
    python morgan_stanley_xlsx.py output.xlsx --lang en
"""

import sys
import argparse
from copy import copy
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, Reference
)
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph as ChartParagraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

# ============================================================
# 1. CONSTANTS & THEME - Enhanced Semantic Color Scheme
# ============================================================

# --- Primary palette (from MS original model) ---
DEEP_NAVY      = "1F3864"   # Primary title background
MED_BLUE       = "2E75B6"   # Subtitle / header background
NAVY_OLD       = "003087"   # Legacy (kept for tab colors)
GOLD           = "C9A227"   # Accent / ticker
WHITE          = "FFFFFF"
BLACK          = "000000"

# --- Semantic fills ---
FILL_INPUT_BG      = "FFF2CC"   # Input assumption background (light yellow)
FILL_UFCF_BG       = "E2EFDA"   # Key calculation row (light green)
FILL_SUMMARY_BG    = "D9E1F2"   # Summary row (light blue-purple)
FILL_CAPEX_BG      = "FCE4D6"   # CapEx / delta NWC row (light orange)
FILL_BEAR_BG       = "FCE4D6"   # Bear scenario cell background
FILL_BASE_BG       = "E2EFDA"   # Base scenario cell background
FILL_BULL_BG       = "DDEEFF"   # Bull scenario cell background
FILL_ALT_ROW       = "F2F2F2"   # Alternating row (light gray)
FILL_CROSS_HL      = "FFFF00"   # Base cross-point highlight (pure yellow)

# --- Scenario identity colors ---
BEAR_COLOR     = "843C0C"   # Brown-orange for bear scenario
BASE_COLOR     = "375623"   # Deep green for base scenario
BULL_COLOR     = "1F3864"   # Deep navy for bull scenario

# --- Font colors ---
BLUE_INPUT     = "2F75B6"   # Input assumption font
RED_NEGATIVE   = "C00000"   # Negative number font
GRAY_NOTE      = "595959"   # Note / description text
DARK_GRAY      = "404040"
GREEN_ACTUAL   = "006100"
LIGHT_GRAY     = "F2F2F2"
MED_GRAY       = "D9D9D9"

# --- Reusable style objects ---
FONT_EN = "Arial"
FONT_ZH = "Microsoft YaHei"

FONT_TITLE      = Font(name=FONT_EN, size=18, bold=True, color=DEEP_NAVY)
FONT_SUBTITLE   = Font(name=FONT_EN, size=12, bold=False, italic=True, color=DARK_GRAY)
FONT_HEADER     = Font(name=FONT_EN, size=10, bold=True, color=WHITE)
FONT_SECTION    = Font(name=FONT_EN, size=10, bold=True, color=DEEP_NAVY)
FONT_LABEL      = Font(name=FONT_EN, size=9, color=BLACK)
FONT_INPUT      = Font(name=FONT_EN, size=9, color=BLUE_INPUT)
FONT_ACTUAL     = Font(name=FONT_EN, size=9, color=GREEN_ACTUAL)
FONT_FORMULA    = Font(name=FONT_EN, size=9, color=BLACK)
FONT_OUTPUT     = Font(name=FONT_EN, size=9, bold=True, color=BLACK)
FONT_ALERT      = Font(name=FONT_EN, size=9, color=RED_NEGATIVE)
FONT_GOLD_HEADER = Font(name=FONT_EN, size=10, bold=True, color=DEEP_NAVY)
FONT_SMALL      = Font(name=FONT_EN, size=8, color=GRAY_NOTE)
FONT_NOTE       = Font(name=FONT_EN, size=9, color=GRAY_NOTE)

# --- Fills ---
FILL_DEEP_NAVY  = PatternFill("solid", fgColor=DEEP_NAVY)
FILL_MED_BLUE   = PatternFill("solid", fgColor=MED_BLUE)
FILL_NAVY       = PatternFill("solid", fgColor=NAVY_OLD)
FILL_LIGHT_BLUE = PatternFill("solid", fgColor="E8EEF4")
FILL_LIGHT_GRAY = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_WHITE      = PatternFill("solid", fgColor=WHITE)
FILL_YELLOW     = PatternFill("solid", fgColor=FILL_INPUT_BG)
FILL_GOLD       = PatternFill("solid", fgColor=GOLD)
FILL_MED_GRAY   = PatternFill("solid", fgColor=MED_GRAY)
FILL_UFCF       = PatternFill("solid", fgColor=FILL_UFCF_BG)
FILL_SUMMARY    = PatternFill("solid", fgColor=FILL_SUMMARY_BG)
FILL_CAPEX      = PatternFill("solid", fgColor=FILL_CAPEX_BG)
FILL_ALT        = PatternFill("solid", fgColor=FILL_ALT_ROW)
FILL_CROSS      = PatternFill("solid", fgColor=FILL_CROSS_HL)
FILL_BEAR_CELL  = PatternFill("solid", fgColor=FILL_BEAR_BG)
FILL_BASE_CELL  = PatternFill("solid", fgColor=FILL_BASE_BG)
FILL_BULL_CELL  = PatternFill("solid", fgColor=FILL_BULL_BG)

# --- Chart palette (MS 8-color) ---
CHART_PALETTE = [
    '1F3864', '2E75B6', 'C8A951', '00AF50',
    'E37C2B', 'B91C1C', '6B4C9A', '4A90D9',
]

# --- Chart style constants ---
CHART_TITLE_SIZE = 1200      # 10pt
CHART_LEGEND_SIZE = 900       # 8pt
CHART_LABEL_SIZE = 800        # 7pt
CHART_GRID_COLOR = "E0E0E0"
CHART_AXIS_COLOR = "4A4A4A"
CHART_WIDTH_CM = 18
CHART_HEIGHT_CM = 10
CHART_GAP_WIDTH = 80
CHART_LINE_WIDTH = 25000      # 2pt

# --- Alignments ---
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_TITLE  = Alignment(horizontal="left", vertical="center")

# --- Borders ---
BORDER_THIN = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
BORDER_BOTTOM_DOUBLE = Border(bottom=Side(style="double", color=BLACK))
BORDER_BOTTOM_THICK = Border(bottom=Side(style="medium", color=BLACK))
BORDER_HEADER = Border(
    left=Side(style="thin", color=DEEP_NAVY),
    right=Side(style="thin", color=DEEP_NAVY),
    top=Side(style="thin", color=DEEP_NAVY),
    bottom=Side(style="medium", color=DEEP_NAVY),
)

# --- Number formats ---
NUM_PCT       = '0.0%'
NUM_PCT2      = '0.00%'
NUM_CURRENCY  = '#,##0'
NUM_CURRENCY_DEC = '#,##0.0'
NUM_MULT      = '0.0"x"'
NUM_MULT2     = '0.00"x"'
NUM_USD       = '$#,##0'
NUM_USD_DEC   = '$#,##0.0'
NUM_USD_M     = '$#,##0.0"M"'
NUM_USD_B     = '$#,##0.0"B"'

# ============================================================
# 2. TRANSLATIONS
# ============================================================

LABELS = {
    "zh": {
        "cover_title": "估值模型",
        "cover_prepared": "分析团队",
        "cover_date": "报告日期",
        "cover_currency": "货币单位",
        "cover_shares": "流通股数（百万）",
        "scenario_summary": "情景假设摘要",
        "scenario_bear": "熊市情景",
        "scenario_base": "基准情景",
        "scenario_bull": "牛市情景",
        "revenue_cagr": "收入CAGR",
        "wacc": "WACC",
        "terminal_growth": "终值增长率",
        "exit_multiple": "退出倍数",
        "target_price": "目标价",
        "implied_ev_rev": "隐含EV/Revenue",
        "implied_ev_ebitda": "隐含EV/EBITDA",
        "valuation_summary": "跨情景估值汇总",
        "enterprise_value": "企业价值",
        "equity_value": "股权价值",
        "per_share_value": "每股价值",
        "net_debt": "净债务",
        "historical": "历史数据",
        "forecast": "预测期",
        "total_revenue": "总收入",
        "ebitda": "EBITDA",
        "ebitda_margin": "EBITDA利润率",
        "da": "折旧与摊销",
        "da_pct": "D&A占收入%",
        "ebit": "EBIT",
        "nopat": "NOPAT",
        "tax_rate": "税率",
        "capex": "资本支出",
        "capex_pct": "CapEx占收入%",
        "delta_nwc": "营运资金变动",
        "nwc_pct": "ΔNWC占收入变化%",
        "ufcf": "自由现金流(UFCF)",
        "pv_fcf": "FCF现值",
        "terminal_value": "终值",
        "tv_ggm": "GGM终值",
        "tv_exit": "退出倍数终值",
        "tv_avg": "平均终值",
        "pv_tv": "终值现值",
        "total_pv": "总现值",
        "sum_pv_fcf": "FCF现值合计",
        "wacc_decomp": "WACC拆解",
        "risk_free_rate": "无风险利率",
        "equity_risk_premium": "股权风险溢价",
        "beta": "Beta",
        "size_premium": "规模溢价",
        "country_risk": "国家风险溢价",
        "cost_of_equity": "股权成本(Ke)",
        "cost_of_debt": "债务成本(Kd)",
        "kd_after_tax": "Kd*(1-T)",
        "tax_rate_w": "税率",
        "equity_weight": "股权权重 E/(D+E)",
        "debt_weight": "债务权重 D/(D+E)",
        "wacc_label": "WACC",
        "sensitivity": "敏感性分析",
        "sens_wacc_tgr_ev": "WACC × 终值增长率 → 企业价值",
        "sens_wacc_tgr_ps": "WACC × 终值增长率 → 每股价值",
        "sens_wacc_mult_ev": "WACC × 退出倍数 → 企业价值",
        "scenario_comparison": "情景对比汇总",
        "comps": "可比公司分析",
        "comp_name": "公司名称",
        "comp_ticker": "代码",
        "comp_mcap": "市值",
        "comp_ev": "企业价值",
        "comp_ev_rev_ltm": "EV/Revenue LTM",
        "comp_ev_rev_ntm": "EV/Revenue NTM",
        "comp_ev_ebitda_ltm": "EV/EBITDA LTM",
        "comp_ev_ebitda_ntm": "EV/EBITDA NTM",
        "comp_cagr": "3Y CAGR",
        "comp_note": "备注",
        "comp_median": "中位数",
        "comp_mean": "平均值",
        "key_assumptions": "关键假设",
        "core_assumptions": "核心假设",
        "disclaimer": "免责声明：本模型仅供研究参考，不构成投资建议。",
        "fy25_base": "FY25基准",
        "growth_rate": "增速",
        "input_assumption": "输入假设",
        "calculated": "计算值",
        "actual": "实际值",
        "business_overview": "业务概览",
        "segment_name": "业务板块",
        "segment_desc": "描述",
        "segment_rev_est": "营收估计",
        "revenue_yoy": "营收同比增速",
        "discount_period": "折现期",
        "shares_outstanding": "流通股本",
        "key_inputs": "关键输入参数",
        "wacc_source": "WACC来源",
        # V7: SOTP
        "sotp_title": "分部估值 / SOTP Valuation",
        "sotp_segment": "业务板块",
        "sotp_ltm_revenue": "LTM收入",
        "sotp_ltm_ebitda": "LTM EBITDA",
        "sotp_method": "估值方法",
        "sotp_multiple": "估值倍数",
        "sotp_ev": "企业价值EV",
        "sotp_note": "备注",
        "sotp_total_ev": "合计企业价值",
        "sotp_less_net_debt": "减：净债务",
        "sotp_less_minority": "减：少数股东权益",
        "sotp_equity_value": "股权价值",
        "sotp_shares": "流通股本（百万）",
        "sotp_per_share": "每股价值",
        "sotp_ev_ebitda_method": "EV/EBITDA",
        "sotp_ev_rev_method": "EV/Revenue",
        "sotp_pe_method": "P/E",
        "sotp_pbook_method": "P/B",
        # V7: KPI
        "kpi_title": "关键运营指标 / Operational KPIs",
        "kpi_metric": "指标",
        "kpi_unit": "单位",
        "kpi_hist_prefix": "历史",
        "kpi_fcst_prefix": "预测",
        # V7: PE Band
        "pe_band_title": "历史PE Band / Historical PE Band",
        "pe_band_current": "当前PE",
        "pe_band_forward": "前瞻PE(NTM)",
        "pe_band_median": "历史中位数",
        "pe_band_period": "期间",
        "pe_band_high": "最高",
        "pe_band_low": "最低",
        "pe_band_close": "收盘价",
        "pe_band_eps": "EPS",
        "pe_band_note": "注：黄色虚线为中位数PE，红色虚线为当前PE",
        # V7: Multi-market comps
        "comp_market": "市场",
        "comp_market_us": "美股",
        "comp_market_hk": "港股",
        "comp_market_a": "A股",
        "comp_market_other": "其他市场",
        "comp_sub_median": "分组中位数",
    },
    "en": {
        "cover_title": "Valuation Model",
        "cover_prepared": "Prepared by",
        "cover_date": "Report Date",
        "cover_currency": "Currency",
        "cover_shares": "Shares Outstanding (M)",
        "scenario_summary": "Scenario Summary",
        "scenario_bear": "Bear Case",
        "scenario_base": "Base Case",
        "scenario_bull": "Bull Case",
        "revenue_cagr": "Revenue CAGR",
        "wacc": "WACC",
        "terminal_growth": "Terminal Growth Rate",
        "exit_multiple": "Exit Multiple",
        "target_price": "Target Price",
        "implied_ev_rev": "Implied EV/Revenue",
        "implied_ev_ebitda": "Implied EV/EBITDA",
        "valuation_summary": "Cross-Scenario Valuation Summary",
        "enterprise_value": "Enterprise Value",
        "equity_value": "Equity Value",
        "per_share_value": "Per Share Value",
        "net_debt": "Net Debt",
        "historical": "Historical",
        "forecast": "Forecast",
        "total_revenue": "Total Revenue",
        "ebitda": "EBITDA",
        "ebitda_margin": "EBITDA Margin",
        "da": "Depreciation & Amortization",
        "da_pct": "D&A % of Revenue",
        "ebit": "EBIT",
        "nopat": "NOPAT",
        "tax_rate": "Tax Rate",
        "capex": "Capital Expenditure",
        "capex_pct": "CapEx % of Revenue",
        "delta_nwc": "Change in NWC",
        "nwc_pct": "ΔNWC % of Rev Change",
        "ufcf": "Unlevered Free Cash Flow",
        "pv_fcf": "PV of FCF",
        "terminal_value": "Terminal Value",
        "tv_ggm": "GGM Terminal Value",
        "tv_exit": "Exit Multiple TV",
        "tv_avg": "Average TV",
        "pv_tv": "PV of Terminal Value",
        "total_pv": "Total Present Value",
        "sum_pv_fcf": "Sum of PV FCF",
        "wacc_decomp": "WACC Decomposition",
        "risk_free_rate": "Risk-Free Rate",
        "equity_risk_premium": "Equity Risk Premium",
        "beta": "Beta",
        "size_premium": "Size Premium",
        "country_risk": "Country Risk Premium",
        "cost_of_equity": "Cost of Equity (Ke)",
        "cost_of_debt": "Cost of Debt (Kd)",
        "kd_after_tax": "Kd*(1-T)",
        "tax_rate_w": "Tax Rate",
        "equity_weight": "Equity Weight E/(D+E)",
        "debt_weight": "Debt Weight D/(D+E)",
        "wacc_label": "WACC",
        "sensitivity": "Sensitivity Analysis",
        "sens_wacc_tgr_ev": "WACC × Terminal Growth → Enterprise Value",
        "sens_wacc_tgr_ps": "WACC × Terminal Growth → Per Share Value",
        "sens_wacc_mult_ev": "WACC × Exit Multiple → Enterprise Value",
        "scenario_comparison": "Scenario Comparison",
        "comps": "Comparable Companies",
        "comp_name": "Company",
        "comp_ticker": "Ticker",
        "comp_mcap": "Market Cap",
        "comp_ev": "Enterprise Value",
        "comp_ev_rev_ltm": "EV/Revenue LTM",
        "comp_ev_rev_ntm": "EV/Revenue NTM",
        "comp_ev_ebitda_ltm": "EV/EBITDA LTM",
        "comp_ev_ebitda_ntm": "EV/EBITDA NTM",
        "comp_cagr": "3Y CAGR",
        "comp_note": "Note",
        "comp_median": "Median",
        "comp_mean": "Mean",
        "key_assumptions": "Key Assumptions",
        "core_assumptions": "Core Assumptions",
        "disclaimer": "Disclaimer: This model is for research purposes only and does not constitute investment advice.",
        "fy25_base": "FY25 Base",
        "growth_rate": "Growth Rate",
        "input_assumption": "Input Assumption",
        "calculated": "Calculated",
        "actual": "Actual",
        "business_overview": "Business Overview",
        "segment_name": "Business Segment",
        "segment_desc": "Description",
        "segment_rev_est": "Revenue Est.",
        "revenue_yoy": "Revenue YoY Growth",
        "discount_period": "Discount Period",
        "shares_outstanding": "Shares Outstanding",
        "key_inputs": "Key Inputs",
        "wacc_source": "WACC Source",
        # V7: SOTP
        "sotp_title": "SOTP Valuation / Sum-of-the-Parts",
        "sotp_segment": "Segment",
        "sotp_ltm_revenue": "LTM Revenue",
        "sotp_ltm_ebitda": "LTM EBITDA",
        "sotp_method": "Valuation Method",
        "sotp_multiple": "Multiple",
        "sotp_ev": "Enterprise Value",
        "sotp_note": "Note",
        "sotp_total_ev": "Total EV",
        "sotp_less_net_debt": "Less: Net Debt",
        "sotp_less_minority": "Less: Minority Interest",
        "sotp_equity_value": "Equity Value",
        "sotp_shares": "Shares Outstanding (M)",
        "sotp_per_share": "Per Share Value",
        "sotp_ev_ebitda_method": "EV/EBITDA",
        "sotp_ev_rev_method": "EV/Revenue",
        "sotp_pe_method": "P/E",
        "sotp_pbook_method": "P/B",
        # V7: KPI
        "kpi_title": "Operational KPIs",
        "kpi_metric": "Metric",
        "kpi_unit": "Unit",
        "kpi_hist_prefix": "Hist.",
        "kpi_fcst_prefix": "Fcast.",
        # V7: PE Band
        "pe_band_title": "Historical PE Band",
        "pe_band_current": "Current PE",
        "pe_band_forward": "Forward PE (NTM)",
        "pe_band_median": "Historical Median",
        "pe_band_period": "Period",
        "pe_band_high": "High",
        "pe_band_low": "Low",
        "pe_band_close": "Close",
        "pe_band_eps": "EPS",
        "pe_band_note": "Note: Yellow dashed = median PE, Red dashed = current PE",
        # V7: Multi-market comps
        "comp_market": "Market",
        "comp_market_us": "US Market",
        "comp_market_hk": "HK Market",
        "comp_market_a": "A-Share Market",
        "comp_market_other": "Other Markets",
        "comp_sub_median": "Sub-Median",
    },
}


def t(key: str, lang: str = "zh") -> str:
    """Get translated label."""
    return LABELS.get(lang, LABELS["zh"]).get(key, key)


# ============================================================
# 3. HELPER FUNCTIONS
# ============================================================

def _font_for_lang(lang: str) -> str:
    return FONT_ZH if lang == "zh" else FONT_EN


def _clone_font(base: Font, name: str = None) -> Font:
    f = copy(base)
    if name:
        f.name = name
    return f


def set_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=None, number_format=None):
    """Write a value to a cell with styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def write_header_row(ws, row, col_start, labels_list, lang="zh"):
    """Write a deep navy header row with white text."""
    fn = _font_for_lang(lang)
    for i, label in enumerate(labels_list):
        c = col_start + i
        set_cell(ws, row, c, label,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_DEEP_NAVY,
                 alignment=ALIGN_CENTER,
                 border=BORDER_HEADER)


def write_sub_header_row(ws, row, col_start, labels_list, lang="zh"):
    """Write a medium blue sub-header row with white text."""
    fn = _font_for_lang(lang)
    for i, label in enumerate(labels_list):
        c = col_start + i
        set_cell(ws, row, c, label,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_MED_BLUE,
                 alignment=ALIGN_CENTER,
                 border=BORDER_HEADER)


def write_section_header(ws, row, col, label, lang="zh"):
    """Write a section header label."""
    fn = _font_for_lang(lang)
    set_cell(ws, row, col, label,
             font=_clone_font(FONT_SECTION, fn),
             fill=FILL_LIGHT_BLUE,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)


def apply_border_range(ws, min_row, max_row, min_col, max_col, border=BORDER_THIN):
    """Apply border to a range of cells."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


def set_col_widths(ws, widths_dict):
    """Set column widths. widths_dict = {col_index: width}."""
    for col, w in widths_dict.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def currency_fmt(cur: str) -> str:
    if cur == "CNY":
        return "\xA5#,##0"
    return "$#,##0"


def currency_dec_fmt(cur: str) -> str:
    if cur == "CNY":
        return "\xA5#,##0.0"
    return "$#,##0.0"


def currency_symbol(cur: str) -> str:
    return "\xA5" if cur == "CNY" else "$"


def _scenario_fill(scenario: str) -> PatternFill:
    """Return scenario-specific cell background fill."""
    if scenario == "bear":
        return FILL_BEAR_CELL
    elif scenario == "base":
        return FILL_BASE_CELL
    else:
        return FILL_BULL_CELL


def _scenario_font_color(scenario: str) -> str:
    """Return scenario identity font color."""
    if scenario == "bear":
        return BEAR_COLOR
    elif scenario == "base":
        return BASE_COLOR
    else:
        return BULL_COLOR


def _alt_row_fill(row_idx: int, base_fill: PatternFill = None) -> PatternFill:
    """Return alternating row fill."""
    if row_idx % 2 == 0:
        return FILL_ALT
    return base_fill or FILL_WHITE


def _ensure_margin_length(margins, n_forecast):
    """Ensure EBITDA margin array length matches forecast years."""
    if len(margins) < n_forecast:
        last_val = margins[-1] if margins else 0.20
        return margins + [last_val] * (n_forecast - len(margins))
    return margins[:n_forecast]


# ============================================================
# 4. COVER SHEET (Enhanced)
# ============================================================

def write_cover_sheet(wb: Workbook, data: dict, lang: str = "zh"):
    """Sheet 1: Cover with business overview, scenario summary, and valuation summary."""
    ws = wb.active
    ws.title = t("cover_title", lang) if lang == "zh" else "Cover"
    ws.sheet_properties.tabColor = DEEP_NAVY

    fn = _font_for_lang(lang)
    cur = data.get("currency", "USD")
    cf = currency_fmt(cur)
    cf_dec = currency_dec_fmt(cur)

    # Column widths
    set_col_widths(ws, {1: 3, 2: 24, 3: 20, 4: 20, 5: 20, 6: 20, 7: 3})

    # --- Title Block ---
    r = 2
    set_cell(ws, r, 2, data.get("company_name", "Company Name"),
             font=Font(name=fn, size=20, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_TITLE)
    r = 3
    set_cell(ws, r, 2, data.get("ticker", "TICKER"),
             font=Font(name=fn, size=14, bold=True, color=GOLD),
             alignment=ALIGN_TITLE)
    r = 4
    set_cell(ws, r, 2, data.get("cover_subtitle", t("cover_title", lang)),
             font=Font(name=fn, size=12, italic=True, color=DARK_GRAY),
             alignment=ALIGN_TITLE)

    # --- Key Info ---
    r = 6
    info_pairs = [
        (t("cover_prepared", lang), data.get("analyst", "")),
        (t("cover_date", lang), data.get("report_date", "")),
        (t("cover_currency", lang), cur),
        (t("cover_shares", lang), data.get("shares_outstanding", "")),
    ]
    for label, val in info_pairs:
        set_cell(ws, r, 2, label, font=_clone_font(FONT_LABEL, fn), alignment=ALIGN_LEFT)
        set_cell(ws, r, 3, val, font=_clone_font(FONT_INPUT, fn), alignment=ALIGN_LEFT)
        r += 1

    # --- Business Overview Section ---
    business_overview = data.get("business_overview", [])
    if business_overview:
        r += 1
        set_cell(ws, r, 2, t("business_overview", lang),
                 font=Font(name=fn, size=12, bold=True, color=DEEP_NAVY),
                 alignment=ALIGN_LEFT)
        r += 1
        bo_headers = [t("segment_name", lang), t("segment_desc", lang), t("segment_rev_est", lang)]
        write_sub_header_row(ws, r, 2, bo_headers, lang)
        r += 1
        bo_start_row = r
        for bi, biz in enumerate(business_overview):
            row_fill = _alt_row_fill(bi)
            set_cell(ws, r, 2, biz.get("name", ""),
                     font=_clone_font(FONT_LABEL, fn),
                     fill=row_fill, alignment=ALIGN_LEFT, border=BORDER_THIN)
            set_cell(ws, r, 3, biz.get("description", ""),
                     font=_clone_font(FONT_NOTE, fn),
                     fill=row_fill, alignment=ALIGN_LEFT, border=BORDER_THIN)
            rev_est = biz.get("revenue_estimate", "")
            set_cell(ws, r, 4, rev_est,
                     font=_clone_font(FONT_INPUT, fn),
                     fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                     number_format=cf if isinstance(rev_est, (int, float)) else None)
            r += 1
        r += 1

    # --- Scenario Summary Table (with core assumptions column) ---
    set_cell(ws, r, 2, t("scenario_summary", lang),
             font=Font(name=fn, size=12, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)
    r += 1

    scenarios_order = ["bear", "base", "bull"]
    scenario_labels = [t("scenario_bear", lang), t("scenario_base", lang), t("scenario_bull", lang)]
    sheet_names = [_scenario_sheet_name(s, lang) for s in scenarios_order]

    # Header: Key Assumptions | Bear | Base | Bull
    headers = [t("key_assumptions", lang)] + scenario_labels
    write_header_row(ws, r, 2, headers, lang)
    r += 1

    summary_rows = [
        (t("revenue_cagr", lang), "cagr"),
        (t("wacc", lang), "wacc"),
        (t("terminal_growth", lang), "tgr"),
        (t("exit_multiple", lang), "exit_mult"),
        (t("enterprise_value", lang), "ev"),
        (t("equity_value", lang), "equity"),
        (t("per_share_value", lang), "ps"),
        (t("implied_ev_rev", lang), "ev_rev"),
        (t("implied_ev_ebitda", lang), "ev_ebitda"),
    ]

    cover_summary_start_row = r
    for idx, (label, key) in enumerate(summary_rows):
        row_fill = _alt_row_fill(idx)
        set_cell(ws, r, 2, label,
                 font=_clone_font(FONT_LABEL, fn),
                 fill=row_fill,
                 alignment=ALIGN_LEFT,
                 border=BORDER_THIN)
        for si, sname in enumerate(sheet_names):
            col = 3 + si
            scenario = scenarios_order[si]
            cell_fill = _scenario_fill(scenario)
            set_cell(ws, r, col, None,
                     font=_clone_font(FONT_FORMULA, fn),
                     fill=cell_fill,
                     alignment=ALIGN_CENTER,
                     border=BORDER_THIN)
        r += 1

    # --- Valuation Summary ---
    r += 1
    set_cell(ws, r, 2, t("valuation_summary", lang),
             font=Font(name=fn, size=12, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)
    r += 1
    val_headers = [""] + scenario_labels
    write_header_row(ws, r, 2, val_headers, lang)
    r += 1
    val_items = [
        t("enterprise_value", lang),
        t("net_debt", lang),
        t("equity_value", lang),
        t("per_share_value", lang),
    ]
    val_summary_start_row = r
    for vi, item in enumerate(val_items):
        set_cell(ws, r, 2, item,
                 font=_clone_font(FONT_LABEL, fn),
                 alignment=ALIGN_LEFT,
                 border=BORDER_THIN)
        for si in range(3):
            scenario = scenarios_order[si]
            cell_fill = _scenario_fill(scenario)
            set_cell(ws, r, 3 + si, None,
                     font=_clone_font(FONT_FORMULA, fn),
                     fill=cell_fill,
                     alignment=ALIGN_CENTER,
                     border=BORDER_THIN)
        r += 1

    # --- Disclaimer ---
    r += 2
    set_cell(ws, r, 2, t("disclaimer", lang),
             font=_clone_font(FONT_SMALL, fn),
             alignment=ALIGN_LEFT)

    # Store metadata for second-pass formula writing
    ws._ms_cover = {
        "summary_start_row": cover_summary_start_row,
        "summary_rows_keys": [k for _, k in summary_rows],
        "val_summary_start_row": val_summary_start_row,
        "scenario_sheet_names": sheet_names,
    }

    return ws


def fill_cover_formulas(wb: Workbook, data: dict, dcf_layouts: dict, lang: str = "zh"):
    """Second pass: fill cover sheet formulas referencing DCF output cells."""
    ws_cover = None
    for ws in wb.worksheets:
        if hasattr(ws, "_ms_cover"):
            ws_cover = ws
            break
    if ws_cover is None:
        return

    meta = ws_cover._ms_cover
    scenarios_order = ["bear", "base", "bull"]
    fn = _font_for_lang(lang)
    cur = data.get("currency", "USD")

    # Fill scenario summary formulas
    r = meta["summary_start_row"]
    for key in meta["summary_rows_keys"]:
        for si, scenario in enumerate(scenarios_order):
            layout = dcf_layouts[scenario]
            col = 3 + si
            cell_ref = layout.get(key)
            if cell_ref:
                formula = f"='{layout['sheet_name']}'!{cell_ref}"
                cell = ws_cover.cell(row=r, column=col)
                cell.value = formula
                cell.font = _clone_font(FONT_FORMULA, fn)
                if key in ("wacc", "tgr", "cagr"):
                    cell.number_format = NUM_PCT
                elif key in ("exit_mult", "ev_rev", "ev_ebitda"):
                    cell.number_format = NUM_MULT
                elif key == "ps":
                    cell.number_format = currency_dec_fmt(cur)
                elif key in ("ev", "equity"):
                    cell.number_format = currency_fmt(cur)
        r += 1

    # Fill valuation summary formulas
    r = meta["val_summary_start_row"]
    val_keys = ["ev", "net_debt", "equity", "ps"]
    for key in val_keys:
        for si, scenario in enumerate(scenarios_order):
            layout = dcf_layouts[scenario]
            col = 3 + si
            cell_ref = layout.get(key)
            if cell_ref:
                formula = f"='{layout['sheet_name']}'!{cell_ref}"
                cell = ws_cover.cell(row=r, column=col)
                cell.value = formula
                cell.font = _clone_font(FONT_FORMULA, fn)
                if key == "ps":
                    cell.number_format = currency_dec_fmt(cur)
                else:
                    cell.number_format = currency_fmt(cur)
        r += 1

    # Highlight per-share value row with yellow
    for si in range(3):
        ws_cover.cell(row=meta["val_summary_start_row"] + 3, column=3 + si).fill = FILL_CROSS


# ============================================================
# 5. DCF SHEET (Enhanced with YoY, Discount Period, separate D&A/EBIT, WACC linkage)
# ============================================================

def _scenario_sheet_name(scenario: str, lang: str = "zh") -> str:
    names = {
        "bear": ("熊市DCF", "Bear DCF"),
        "base": ("基准DCF", "Base DCF"),
        "bull": ("牛市DCF", "Bull DCF"),
    }
    return names[scenario][0] if lang == "zh" else names[scenario][1]


def write_dcf_sheet(wb: Workbook, data: dict, scenario: str, lang: str = "zh",
                    wacc_sheet_name: str = None, wacc_result_row: int = None):
    """Write a complete DCF sheet for one scenario. Returns layout dict with key cell refs.

    Enhanced V6:
    - Row 1: Title
    - Row 2: Year headers
    - Row 3: Section - Revenue
    - Row 4..4+n_seg-1: Segment revenues
    - Row 4+n_seg: Total Revenue
    - Row 4+n_seg+1: Revenue YoY Growth %
    - Row 4+n_seg+2: Section - Profitability
    - Row 4+n_seg+3: EBITDA Margin (input)
    - Row 4+n_seg+4: EBITDA (calculated)
    - Row 4+n_seg+5: D&A (calculated)
    - Row 4+n_seg+6: EBIT (calculated)
    - Row 4+n_seg+7: NOPAT (calculated)
    - Row 4+n_seg+8: Section - Free Cash Flow
    - Row 4+n_seg+9: CapEx (calculated)
    - Row 4+n_seg+10: Change in NWC (calculated)
    - Row 4+n_seg+11: UFCF (calculated)
    - Row 4+n_seg+12: Section - DCF Valuation
    - Row 4+n_seg+13: Discount Period
    - Row 4+n_seg+14: PV of FCF
    - separator
    - Section: Terminal Value
    - Sum PV FCF, TGR (input), Exit Multiple (input), WACC (from WACC sheet)
    - GGM TV, Exit Multiple TV (separate), Average TV, PV of TV
    - EV, Net Debt, Equity Value, Per Share Value
    - Implied EV/Rev, Implied EV/EBITDA, Revenue CAGR
    """
    sheet_name = _scenario_sheet_name(scenario, lang)
    ws = wb.create_sheet(title=sheet_name)

    fn = _font_for_lang(lang)
    cur = data.get("currency", "USD")
    cf = currency_fmt(cur)
    cf_dec = currency_dec_fmt(cur)

    sc = data["scenarios"][scenario]
    hist = data.get("historical", {})
    hist_years = hist.get("years", [])
    forecast_years = data.get("forecast_years", [])
    n_hist = len(hist_years)
    n_forecast = len(forecast_years)
    n_total = n_hist + n_forecast

    segments = sc.get("segments", {})
    seg_names = list(segments.keys())
    n_seg = len(seg_names)

    # --- Validate EBITDA margin length ---
    ebitda_margins = _ensure_margin_length(sc.get("ebitda_margin", [0.20] * n_forecast), n_forecast)

    col_start = 3  # first data column

    # Column widths
    set_col_widths(ws, {1: 2, 2: 28})
    for i in range(n_total):
        ws.column_dimensions[get_column_letter(col_start + i)].width = 14

    # Scenario-specific tab color
    tab_colors = {"bear": BEAR_COLOR, "base": BASE_COLOR, "bull": BULL_COLOR}
    ws.sheet_properties.tabColor = tab_colors.get(scenario, DEEP_NAVY)

    # --- Row 1: Title ---
    r = 1
    scenario_label = sc.get("label_zh" if lang == "zh" else "label_en", scenario.title())
    scenario_color = _scenario_font_color(scenario)
    set_cell(ws, r, 2, scenario_label,
             font=Font(name=fn, size=14, bold=True, color=scenario_color),
             alignment=ALIGN_LEFT)

    # --- Row 2: Year headers ---
    r = 2
    all_years = hist_years + forecast_years
    headers = [""] + all_years
    write_header_row(ws, r, 2, headers, lang)

    # --- Row 3: Section - Revenue ---
    r = 3
    write_section_header(ws, r, 2, t("total_revenue", lang), lang)

    # --- Rows 4 to 4+n_seg-1: Segment revenues ---
    seg_start_row = 4
    for si, seg_name in enumerate(seg_names):
        r = seg_start_row + si
        seg_data = segments[seg_name]
        row_fill = _alt_row_fill(si)
        set_cell(ws, r, 2, f"  {seg_name}",
                 font=_clone_font(FONT_LABEL, fn),
                 fill=row_fill,
                 alignment=ALIGN_LEFT,
                 border=BORDER_THIN)

        # Historical data if available
        hist_segs = hist.get("segments", {})
        hist_rev = hist_segs.get(seg_name, {}).get("revenue", [])
        for hi in range(n_hist):
            if hi < len(hist_rev):
                set_cell(ws, r, col_start + hi, hist_rev[hi],
                         font=_clone_font(FONT_ACTUAL, fn),
                         fill=row_fill,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN,
                         number_format=cf)

        # Forecast years
        growth_rates = seg_data.get("growth_rates", [0] * n_forecast)
        fy25_base = seg_data.get("fy25_base", 0)
        for fi in range(n_forecast):
            col_idx = col_start + n_hist + fi
            prev_col = get_column_letter(col_idx - 1)
            if fi == 0 and fy25_base:
                # First forecast year (FY25E): use fy25_base directly
                formula = f"={fy25_base}"
            elif fi == 0 and n_hist > 0:
                formula = f"={prev_col}{r}*(1+{growth_rates[fi]:.4f})"
            elif fi == 0:
                formula = f"={fy25_base}*(1+{growth_rates[fi]:.4f})"
            else:
                # growth_rates[fi-1] because fi=0 uses base, rates start from fi=1
                gi = fi - 1 if fy25_base else fi
                gi = min(gi, len(growth_rates) - 1)
                formula = f"={prev_col}{r}*(1+{growth_rates[gi]:.4f})"
            set_cell(ws, r, col_idx, formula,
                     font=_clone_font(FONT_FORMULA, fn),
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf)

    # --- Total Revenue Row ---
    total_rev_row = seg_start_row + n_seg
    r = total_rev_row
    set_cell(ws, r, 2, t("total_revenue", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)

    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        seg_refs = "+".join([f"{col_l}{seg_start_row + si}" for si in range(n_seg)])
        formula = f"={seg_refs}"
        set_cell(ws, r, col_idx, formula,
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- Revenue YoY Growth Row ---
    rev_yoy_row = total_rev_row + 1
    r = rev_yoy_row
    set_cell(ws, r, 2, t("revenue_yoy", lang),
             font=_clone_font(FONT_NOTE, fn),
             fill=_alt_row_fill(0),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        if ci == 0:
            set_cell(ws, r, col_idx, None,
                     font=_clone_font(FONT_NOTE, fn),
                     fill=_alt_row_fill(0),
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=NUM_PCT)
        else:
            prev_col = get_column_letter(col_idx - 1)
            formula = f"=IF({prev_col}{total_rev_row}=0,0,({col_l}{total_rev_row}-{prev_col}{total_rev_row})/{prev_col}{total_rev_row})"
            set_cell(ws, r, col_idx, formula,
                     font=_clone_font(FONT_NOTE, fn),
                     fill=_alt_row_fill(0),
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=NUM_PCT)

    # --- Section: Profitability ---
    r = rev_yoy_row + 1
    write_section_header(ws, r, 2, "Profitability", lang)

    # --- EBITDA Margin Row (INPUT - yellow background, blue font) ---
    r += 1
    ebitda_margin_row = r
    set_cell(ws, r, 2, t("ebitda_margin", lang),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for ci in range(n_total):
        col_idx = col_start + ci
        if ci < n_hist:
            # Historical margin (calculated)
            hist_ebitda = hist.get("ebitda", [])
            hist_rev = hist.get("revenue", [])
            if ci < len(hist_ebitda) and ci < len(hist_rev) and hist_rev[ci] != 0:
                set_cell(ws, r, col_idx, hist_ebitda[ci] / hist_rev[ci],
                         font=_clone_font(FONT_ACTUAL, fn),
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN,
                         number_format=NUM_PCT)
            else:
                set_cell(ws, r, col_idx, None,
                         font=_clone_font(FONT_ACTUAL, fn),
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN,
                         number_format=NUM_PCT)
        else:
            fi = ci - n_hist
            margin = ebitda_margins[fi]
            set_cell(ws, r, col_idx, margin,
                     font=_clone_font(FONT_INPUT, fn),
                     fill=FILL_YELLOW,
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=NUM_PCT)

    # --- EBITDA Row (calculated = Revenue * Margin) ---
    r += 1
    ebitda_row = r
    set_cell(ws, r, 2, t("ebitda", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        if ci < n_hist:
            hist_ebitda = hist.get("ebitda", [])
            if ci < len(hist_ebitda):
                set_cell(ws, r, col_idx, hist_ebitda[ci],
                         font=_clone_font(FONT_ACTUAL, fn),
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN,
                         number_format=cf)
        else:
            formula = f"={col_l}{total_rev_row}*{col_l}{ebitda_margin_row}"
            set_cell(ws, r, col_idx, formula,
                     font=_clone_font(FONT_FORMULA, fn),
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf)

    # --- D&A Row (separate, calculated) ---
    r += 1
    da_row = r
    set_cell(ws, r, 2, t("da", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    da_pct = sc.get("da_pct_revenue", 0.04)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        formula = f"={col_l}{total_rev_row}*{da_pct:.4f}"
        set_cell(ws, r, col_idx, formula,
                 font=_clone_font(FONT_FORMULA, fn),
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- EBIT Row (separate, calculated = EBITDA - D&A) ---
    r += 1
    ebit_row = r
    set_cell(ws, r, 2, t("ebit", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        formula = f"={col_l}{ebitda_row}-{col_l}{da_row}"
        set_cell(ws, r, col_idx, formula,
                 font=_clone_font(FONT_FORMULA, fn),
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- NOPAT Row ---
    r += 1
    nopat_row = r
    set_cell(ws, r, 2, t("nopat", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    tax_rate = sc.get("tax_rate", 0.21)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        formula = f"={col_l}{ebit_row}*(1-{tax_rate:.4f})"
        set_cell(ws, r, col_idx, formula,
                 font=_clone_font(FONT_FORMULA, fn),
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- Section: Free Cash Flow ---
    r += 1
    write_section_header(ws, r, 2, t("ufcf", lang), lang)

    # --- CapEx Row (orange background, red font) ---
    r += 1
    capex_row = r
    set_cell(ws, r, 2, t("capex", lang),
             font=Font(name=fn, size=9, color=RED_NEGATIVE),
             fill=FILL_CAPEX,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    capex_pct = sc.get("capex_pct_revenue", 0.06)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        formula = f"={col_l}{total_rev_row}*{capex_pct:.4f}"
        set_cell(ws, r, col_idx, formula,
                 font=Font(name=fn, size=9, color=RED_NEGATIVE),
                 fill=FILL_CAPEX,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- Change in NWC Row (orange background, red font) ---
    r += 1
    nwc_row = r
    set_cell(ws, r, 2, t("delta_nwc", lang),
             font=Font(name=fn, size=9, color=RED_NEGATIVE),
             fill=FILL_CAPEX,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    nwc_pct = sc.get("nwc_pct_rev_change", 0.10)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        if ci == 0:
            formula = f"={col_l}{total_rev_row}*{nwc_pct * 0.5:.4f}"
        else:
            prev_col = get_column_letter(col_idx - 1)
            formula = f"=({col_l}{total_rev_row}-{prev_col}{total_rev_row})*{nwc_pct:.4f}"
        set_cell(ws, r, col_idx, formula,
                 font=Font(name=fn, size=9, color=RED_NEGATIVE),
                 fill=FILL_CAPEX,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- UFCF Row (green background - key calculation) ---
    r += 1
    ufcf_row = r
    set_cell(ws, r, 2, t("ufcf", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_UFCF,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        formula = f"={col_l}{nopat_row}+{col_l}{da_row}-{col_l}{capex_row}-{col_l}{nwc_row}"
        set_cell(ws, r, col_idx, formula,
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_UFCF,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=cf)

    # --- Section: DCF Valuation ---
    r += 1
    write_section_header(ws, r, 2, t("terminal_value", lang), lang)

    # --- Discount Period Row ---
    r += 1
    discount_period_row = r
    set_cell(ws, r, 2, t("discount_period", lang),
             font=_clone_font(FONT_NOTE, fn),
             fill=_alt_row_fill(0),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for ci in range(n_total):
        col_idx = col_start + ci
        if ci < n_hist:
            # Historical years: no discount period
            set_cell(ws, r, col_idx, None,
                     font=_clone_font(FONT_NOTE, fn),
                     fill=_alt_row_fill(0),
                     alignment=ALIGN_CENTER,
                     border=BORDER_THIN)
        else:
            # Forecast years: discount period starts at 0
            fi = ci - n_hist
            set_cell(ws, r, col_idx, fi,
                     font=_clone_font(FONT_NOTE, fn),
                     fill=_alt_row_fill(0),
                     alignment=ALIGN_CENTER,
                     border=BORDER_THIN)

    # --- PV of FCF Row ---
    r += 1
    pv_fcf_row = r
    set_cell(ws, r, 2, t("pv_fcf", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)

    # WACC: compute locally for PV formula, but also write cross-sheet reference
    wacc = sc["wacc"]
    ke_val = (wacc["rf"] + wacc["beta"] * wacc["erp"]
              + wacc.get("size_premium", 0) + wacc.get("country_risk", 0))
    kd = wacc.get("kd", 0.06)
    wacc_tax = wacc.get("tax_rate", 0.21)
    ew = wacc.get("e_weight", 0.9)
    dw = wacc.get("d_weight", 0.1)
    wacc_computed = ke_val * ew + kd * (1 - wacc_tax) * dw

    for ci in range(n_total):
        col_idx = col_start + ci
        col_l = get_column_letter(col_idx)
        if ci < n_hist:
            # Historical years: no PV calculation
            set_cell(ws, r, col_idx, None,
                     font=_clone_font(FONT_FORMULA, fn),
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf)
        else:
            # Forecast years: PV_FCF_t = UFCF_t / (1 + WACC)^(t+1)
            # where t=0 is the first forecast year (year-end discounting)
            fi = ci - n_hist
            year_idx = fi + 1
            formula = f"={col_l}{ufcf_row}/(1+{wacc_computed:.6f})^{year_idx}"
            set_cell(ws, r, col_idx, formula,
                     font=_clone_font(FONT_FORMULA, fn),
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf)

    # --- Separator ---
    r += 1
    set_cell(ws, r, 2, "", border=BORDER_BOTTOM_THICK)

    # --- Key Inputs Section ---
    r += 1
    write_section_header(ws, r, 2, t("key_inputs", lang), lang)

    # Sum of PV FCF
    r += 1
    sum_pv_fcf_row = r
    set_cell(ws, r, 2, t("sum_pv_fcf", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    last_fc_col = get_column_letter(col_start + n_total - 1)
    first_fc_col = get_column_letter(col_start)
    formula = f"=SUM({first_fc_col}{pv_fcf_row}:{last_fc_col}{pv_fcf_row})"
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # Terminal Growth Rate (input - yellow bg)
    r += 1
    tgr_row = r
    set_cell(ws, r, 2, t("terminal_growth", lang),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    set_cell(ws, r, col_start, sc.get("terminal_growth_rate", 0.02),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=NUM_PCT)

    # Exit Multiple (input - yellow bg)
    r += 1
    exit_mult_row = r
    set_cell(ws, r, 2, t("exit_multiple", lang),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    set_cell(ws, r, col_start, sc.get("exit_multiple_ebitda", 8.0),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=NUM_MULT)

    # Shares Outstanding (input - yellow bg)
    r += 1
    shares_row = r
    set_cell(ws, r, 2, t("shares_outstanding", lang),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    shares = data.get("shares_outstanding", 1000)
    set_cell(ws, r, col_start, shares,
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=NUM_CURRENCY)

    # Net Debt (input - yellow bg)
    r += 1
    net_debt_row = r
    set_cell(ws, r, 2, t("net_debt", lang),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    set_cell(ws, r, col_start, sc.get("net_debt", 0),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # WACC (CROSS-SHEET REFERENCE from WACC Decomposition sheet)
    r += 1
    wacc_row = r
    set_cell(ws, r, 2, t("wacc_label", lang),
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_YELLOW,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    if wacc_sheet_name and wacc_result_row is not None:
        # Cross-sheet formula: =WACC拆解!C{row} for bear, D{row} for base, E{row} for bull
        wacc_col_map = {"bear": "C", "base": "D", "bull": "E"}
        wacc_col = wacc_col_map.get(scenario, "D")
        formula = f"='{wacc_sheet_name}'!{wacc_col}{wacc_result_row}"
        set_cell(ws, r, col_start, formula,
                 font=_clone_font(FONT_INPUT, fn),
                 fill=FILL_YELLOW,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT2)
    else:
        # Fallback: use locally computed value
        set_cell(ws, r, col_start, wacc_computed,
                 font=_clone_font(FONT_INPUT, fn),
                 fill=FILL_YELLOW,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT2)

    # WACC source note
    set_cell(ws, r, col_start + 1, t("wacc_source", lang) + " → " + (wacc_sheet_name or "local"),
             font=_clone_font(FONT_SMALL, fn),
             alignment=ALIGN_LEFT)

    # --- Separator ---
    r += 1
    set_cell(ws, r, 2, "", border=BORDER_BOTTOM_THICK)

    # --- Terminal Value Calculations ---
    r += 1
    write_section_header(ws, r, 2, t("terminal_value", lang), lang)

    # GGM Terminal Value (separate display)
    r += 1
    tv_ggm_row = r
    set_cell(ws, r, 2, t("tv_ggm", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = (f"={last_fc_col}{ufcf_row}*(1+{get_column_letter(col_start)}{tgr_row})"
               f"/({get_column_letter(col_start)}{wacc_row}-{get_column_letter(col_start)}{tgr_row})")
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # Exit Multiple TV (separate display)
    r += 1
    tv_exit_row = r
    set_cell(ws, r, 2, t("tv_exit", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"={last_fc_col}{ebitda_row}*{get_column_letter(col_start)}{exit_mult_row}"
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # Average TV
    r += 1
    tv_avg_row = r
    set_cell(ws, r, 2, t("tv_avg", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"=({get_column_letter(col_start)}{tv_ggm_row}+{get_column_letter(col_start)}{tv_exit_row})/2"
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # PV of TV
    r += 1
    pv_tv_row = r
    set_cell(ws, r, 2, t("pv_tv", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = (f"={get_column_letter(col_start)}{tv_avg_row}"
               f"/(1+{get_column_letter(col_start)}{wacc_row})^{n_forecast}")
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # --- Separator ---
    r += 1
    set_cell(ws, r, 2, "", border=BORDER_BOTTOM_THICK)

    # --- Enterprise Value ---
    r += 1
    ev_row = r
    set_cell(ws, r, 2, t("enterprise_value", lang),
             font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"={get_column_letter(col_start)}{sum_pv_fcf_row}+{get_column_letter(col_start)}{pv_tv_row}"
    set_cell(ws, r, col_start, formula,
             font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # --- Equity Value ---
    r += 1
    equity_row = r
    set_cell(ws, r, 2, t("equity_value", lang),
             font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"={get_column_letter(col_start)}{ev_row}-{get_column_letter(col_start)}{net_debt_row}"
    set_cell(ws, r, col_start, formula,
             font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf)

    # --- Per Share Value ---
    r += 1
    ps_row = r
    set_cell(ws, r, 2, t("per_share_value", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"={get_column_letter(col_start)}{equity_row}/{get_column_letter(col_start)}{shares_row}"
    set_cell(ws, r, col_start, formula,
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=cf_dec)

    # --- Implied Multiples ---
    r += 1
    ev_rev_row = r
    set_cell(ws, r, 2, t("implied_ev_rev", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"=IF({last_fc_col}{total_rev_row}=0,0,{get_column_letter(col_start)}{ev_row}/{last_fc_col}{total_rev_row})"
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=NUM_MULT)

    r += 1
    ev_ebitda_row = r
    set_cell(ws, r, 2, t("implied_ev_ebitda", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    formula = f"=IF({last_fc_col}{ebitda_row}=0,0,{get_column_letter(col_start)}{ev_row}/{last_fc_col}{ebitda_row})"
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=NUM_MULT)

    # --- Revenue CAGR (forecast period) ---
    r += 1
    cagr_row = r
    set_cell(ws, r, 2, t("revenue_cagr", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    first_fc_col = get_column_letter(col_start + n_hist)
    formula = f"=({last_fc_col}{total_rev_row}/{first_fc_col}{total_rev_row})^(1/{n_forecast})-1"
    set_cell(ws, r, col_start, formula,
             font=_clone_font(FONT_FORMULA, fn),
             alignment=ALIGN_RIGHT,
             border=BORDER_THIN,
             number_format=NUM_PCT)

    # Build layout dict for cover sheet references
    layout = {
        "sheet_name": sheet_name,
        "wacc": f"{get_column_letter(col_start)}{wacc_row}",
        "tgr": f"{get_column_letter(col_start)}{tgr_row}",
        "exit_mult": f"{get_column_letter(col_start)}{exit_mult_row}",
        "ev": f"{get_column_letter(col_start)}{ev_row}",
        "equity": f"{get_column_letter(col_start)}{equity_row}",
        "ps": f"{get_column_letter(col_start)}{ps_row}",
        "ev_rev": f"{get_column_letter(col_start)}{ev_rev_row}",
        "ev_ebitda": f"{get_column_letter(col_start)}{ev_ebitda_row}",
        "cagr": f"{get_column_letter(col_start)}{cagr_row}",
        "net_debt": f"{get_column_letter(col_start)}{net_debt_row}",
        "total_rev_row": total_rev_row,
        "ebitda_row": ebitda_row,
        "ebitda_margin_row": ebitda_margin_row,
        "ufcf_row": ufcf_row,
        "pv_fcf_row": pv_fcf_row,
        "sum_pv_fcf_row": sum_pv_fcf_row,
        "tv_ggm_row": tv_ggm_row,
        "tv_exit_row": tv_exit_row,
        "tv_avg_row": tv_avg_row,
        "pv_tv_row": pv_tv_row,
        "net_debt_row": f"{get_column_letter(col_start)}{net_debt_row}",
        "shares_row": f"{get_column_letter(col_start)}{shares_row}",
        "n_hist": n_hist,
        "n_forecast": n_forecast,
        "col_start": col_start,
        "last_fc_col": last_fc_col,
        "wacc_computed": wacc_computed,
    }

    # Apply borders to the entire data range
    apply_border_range(ws, 2, r, 2, col_start + n_total - 1, BORDER_THIN)

    # Freeze panes
    ws.freeze_panes = f"C3"

    return layout


# ============================================================
# 6. WACC DECOMPOSITION SHEET (Enhanced with formula-driven Ke/WACC)
# ============================================================

def write_wacc_sheet(wb: Workbook, data: dict, dcf_layouts: dict, lang: str = "zh"):
    """Sheet: WACC Decomposition - three scenarios side by side.

    Enhanced V6:
    - Row 1: Title
    - Row 2: Headers (Label | Bear | Base | Bull)
    - Row 3: Section - Cost of Equity (CAPM)
    - Row 4: Risk-Free Rate (input)
    - Row 5: Equity Risk Premium (input)
    - Row 6: Beta (input)
    - Row 7: Size Premium (input)
    - Row 8: Country Risk Premium (input)
    - Row 9: Ke = Rf + Beta*ERP + Size Premium + Country Risk (FORMULA)
    - separator
    - Row 11: Section - Cost of Debt
    - Row 12: Kd pre-tax (input)
    - Row 13: Tax Rate (input)
    - Row 14: Kd*(1-T) (FORMULA)
    - separator
    - Row 16: Section - Capital Structure
    - Row 17: Equity Weight E/(D+E) (input)
    - Row 18: Debt Weight D/(D+E) = 1 - E Weight (FORMULA)
    - separator
    - Row 20: WACC = Ke*E + Kd(1-T)*D (FORMULA)
    """
    ws = wb.create_sheet(title=t("wacc_decomp", lang))
    ws.sheet_properties.tabColor = GOLD

    fn = _font_for_lang(lang)
    scenarios_order = ["bear", "base", "bull"]
    scenario_labels = [
        t("scenario_bear", lang), t("scenario_base", lang), t("scenario_bull", lang)
    ]

    set_col_widths(ws, {1: 2, 2: 30, 3: 16, 4: 16, 5: 16})

    # Title
    r = 1
    set_cell(ws, r, 2, t("wacc_decomp", lang),
             font=Font(name=fn, size=14, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)

    # Headers
    r = 2
    headers = [t("wacc_decomp", lang)] + scenario_labels
    write_header_row(ws, r, 2, headers, lang)

    # Section: Cost of Equity (CAPM)
    r = 3
    write_section_header(ws, r, 2, t("cost_of_equity", lang) + " (CAPM)", lang)

    components = [
        ("risk_free_rate", "rf", NUM_PCT),
        ("equity_risk_premium", "erp", NUM_PCT),
        ("beta", "beta", '0.00'),
        ("size_premium", "size_premium", NUM_PCT),
        ("country_risk", "country_risk", NUM_PCT),
    ]

    r = 4
    for label_key, param_key, num_fmt in components:
        set_cell(ws, r, 2, t(label_key, lang),
                 font=_clone_font(FONT_LABEL, fn),
                 alignment=ALIGN_LEFT,
                 border=BORDER_THIN)
        for si, scenario in enumerate(scenarios_order):
            wacc_data = data["scenarios"][scenario]["wacc"]
            val = wacc_data.get(param_key, 0)
            set_cell(ws, r, 3 + si, val,
                     font=_clone_font(FONT_INPUT, fn),
                     fill=FILL_YELLOW,
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=num_fmt)
        r += 1

    # Ke = Rf + Beta * ERP + Size Premium + Country Risk (FORMULA)
    ke_row = r
    set_cell(ws, r, 2, t("cost_of_equity", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    rf_row = 4
    erp_row = 5
    beta_row = 6
    sp_row = 7
    cr_row = 8
    for si in range(3):
        col = 3 + si
        col_l = get_column_letter(col)
        formula = f"={col_l}{rf_row}+{col_l}{beta_row}*{col_l}{erp_row}+{col_l}{sp_row}+{col_l}{cr_row}"
        set_cell(ws, r, col, formula,
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT2)
    r += 1

    # Separator
    set_cell(ws, r, 2, "", border=BORDER_BOTTOM_THICK)
    r += 1

    # Section: Cost of Debt
    write_section_header(ws, r, 2, t("cost_of_debt", lang), lang)
    r += 1

    kd_row = r
    set_cell(ws, r, 2, t("cost_of_debt", lang) + " (Pre-tax)",
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for si, scenario in enumerate(scenarios_order):
        val = data["scenarios"][scenario]["wacc"].get("kd", 0.06)
        set_cell(ws, r, 3 + si, val,
                 font=_clone_font(FONT_INPUT, fn),
                 fill=FILL_YELLOW,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
    r += 1

    tax_row = r
    set_cell(ws, r, 2, t("tax_rate_w", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for si, scenario in enumerate(scenarios_order):
        val = data["scenarios"][scenario]["wacc"].get("tax_rate", 0.21)
        set_cell(ws, r, 3 + si, val,
                 font=_clone_font(FONT_INPUT, fn),
                 fill=FILL_YELLOW,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
    r += 1

    # Kd*(1-T) (FORMULA)
    kd_after_tax_row = r
    set_cell(ws, r, 2, t("kd_after_tax", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for si in range(3):
        col = 3 + si
        col_l = get_column_letter(col)
        formula = f"={col_l}{kd_row}*(1-{col_l}{tax_row})"
        set_cell(ws, r, col, formula,
                 font=_clone_font(FONT_FORMULA, fn),
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT2)
    r += 1

    # Separator
    set_cell(ws, r, 2, "", border=BORDER_BOTTOM_THICK)
    r += 1

    # Section: Capital Structure
    write_section_header(ws, r, 2, "Capital Structure", lang)
    r += 1

    ew_row = r
    set_cell(ws, r, 2, t("equity_weight", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for si, scenario in enumerate(scenarios_order):
        val = data["scenarios"][scenario]["wacc"].get("e_weight", 0.9)
        set_cell(ws, r, 3 + si, val,
                 font=_clone_font(FONT_INPUT, fn),
                 fill=FILL_YELLOW,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
    r += 1

    # Debt Weight = 1 - Equity Weight (FORMULA)
    dw_row = r
    set_cell(ws, r, 2, t("debt_weight", lang),
             font=_clone_font(FONT_LABEL, fn),
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for si in range(3):
        col = 3 + si
        col_l = get_column_letter(col)
        formula = f"=1-{col_l}{ew_row}"
        set_cell(ws, r, col, formula,
                 font=_clone_font(FONT_FORMULA, fn),
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
    r += 1

    # Separator
    set_cell(ws, r, 2, "", border=BORDER_BOTTOM_THICK)
    r += 1

    # WACC = Ke*E + Kd(1-T)*D (FORMULA)
    wacc_result_row = r
    set_cell(ws, r, 2, t("wacc_label", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             fill=FILL_CROSS,
             alignment=ALIGN_LEFT,
             border=BORDER_THIN)
    for si in range(3):
        col = 3 + si
        col_l = get_column_letter(col)
        formula = f"={col_l}{ke_row}*{col_l}{ew_row}+{col_l}{kd_after_tax_row}*{col_l}{dw_row}"
        set_cell(ws, r, col, formula,
                 font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
                 fill=FILL_CROSS,
                 alignment=ALIGN_RIGHT,
                 border=BORDER_THIN,
                 number_format=NUM_PCT2)

    # Apply borders
    apply_border_range(ws, 2, r, 2, 5, BORDER_THIN)

    return ws, wacc_result_row


# ============================================================
# 7. SENSITIVITY ANALYSIS SHEET (Enhanced with correct refs & yellow highlight)
# ============================================================

def write_sensitivity_sheet(wb: Workbook, data: dict, dcf_layouts: dict, lang: str = "zh"):
    """Sheet: Sensitivity Analysis with 4 matrices.

    Enhanced V6:
    - Matrix 1: WACC x TGR -> Enterprise Value (formulas)
    - Matrix 2: WACC x TGR -> Per Share Value (formulas)
    - Matrix 3: WACC x Exit Multiple -> Enterprise Value (formulas, correct EBITDA ref)
    - Matrix 4: Scenario Comparison Summary
    - Base cross-point highlighted in yellow
    """
    ws = wb.create_sheet(title=t("sensitivity", lang))
    ws.sheet_properties.tabColor = "4472C4"

    fn = _font_for_lang(lang)
    cur = data.get("currency", "USD")
    cf = currency_fmt(cur)
    cf_dec = currency_dec_fmt(cur)

    set_col_widths(ws, {1: 2, 2: 22})
    for i in range(3, 14):
        ws.column_dimensions[get_column_letter(i)].width = 14

    base_wacc = dcf_layouts["base"]["wacc_computed"]
    base_tgr = data["scenarios"]["base"].get("terminal_growth_rate", 0.02)
    base_exit_mult = data["scenarios"]["base"].get("exit_multiple_ebitda", 8.0)

    # Support data-driven base index positions for correct cross-highlight
    sens_cfg = data.get("sensitivity", {})
    base_wacc_idx = sens_cfg.get("base_wacc_idx", 4)  # default: center of 9 values
    base_tgr_idx = sens_cfg.get("base_tgr_idx", 3)    # default: center of 7 values

    # WACC range: base_wacc centered at base_wacc_idx, step 0.5%
    n_wacc = 9
    wacc_step = 0.005
    wacc_values = [round(base_wacc + (i - base_wacc_idx) * wacc_step, 4) for i in range(n_wacc)]
    # TGR range: base_tgr centered at base_tgr_idx, step 0.5%
    n_tgr = 7
    tgr_step = 0.005
    tgr_values = [round(base_tgr + (i - base_tgr_idx) * tgr_step, 4) for i in range(n_tgr)]
    # Exit Multiple range: base centered, step 2
    n_mult = 5
    base_mult_idx = 2  # center of 5 values
    mult_values = [round(base_exit_mult + (i - base_mult_idx) * 2, 1) for i in range(n_mult)]

    scenarios_order = ["bear", "base", "bull"]
    base_layout = dcf_layouts["base"]
    n_f = base_layout["n_forecast"]

    # References to base DCF sheet
    ufcf_ref = f"'{base_layout['sheet_name']}'!{base_layout['last_fc_col']}{base_layout['ufcf_row']}"
    ebitda_ref = f"'{base_layout['sheet_name']}'!{base_layout['last_fc_col']}{base_layout['ebitda_row']}"
    sum_pv_ref = f"'{base_layout['sheet_name']}'!{get_column_letter(base_layout['col_start'])}{base_layout['sum_pv_fcf_row']}"
    net_debt_ref = f"'{base_layout['sheet_name']}'!{base_layout['net_debt_row']}"

    shares = data.get("shares_outstanding", 1000)

    # --- Matrix 1: WACC x TGR -> Enterprise Value ---
    r = 1
    set_cell(ws, r, 2, t("sens_wacc_tgr_ev", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)
    r = 2
    set_cell(ws, r, 2, f"{t('wacc', lang)} \\ {t('terminal_growth', lang)}",
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY,
             alignment=ALIGN_CENTER,
             border=BORDER_HEADER)
    for ji, tgr in enumerate(tgr_values):
        set_cell(ws, r, 3 + ji, tgr,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_DEEP_NAVY,
                 alignment=ALIGN_CENTER,
                 border=BORDER_HEADER,
                 number_format=NUM_PCT)

    m1_start_row = 3
    # Base WACC cell reference for PV FCF adjustment (row of base_wacc_idx in matrix)
    base_wacc_row = m1_start_row + base_wacc_idx
    base_wacc_cell_ref = f"${get_column_letter(2)}${base_wacc_row}"

    for wi, wacc in enumerate(wacc_values):
        r = m1_start_row + wi
        set_cell(ws, r, 2, wacc,
                 font=_clone_font(FONT_INPUT, fn),
                 alignment=ALIGN_CENTER,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
        for ji, tgr in enumerate(tgr_values):
            wacc_cell = f"${get_column_letter(2)}${r}"
            tgr_cell = f"${get_column_letter(3 + ji)}$2"

            # EV = PV_GGM_TV + PV_Exit_TV (averaged) + adjusted PV_FCF
            # PV_FCF adjustment: scale base PV_FCF by base_wacc/current_wacc ratio
            formula = (
                f"=({ufcf_ref}*(1+{tgr_cell})/({wacc_cell}-{tgr_cell})"
                f"+{ebitda_ref}*{base_exit_mult})"
                f"/2/(1+{wacc_cell})^{n_f}"
                f"+{sum_pv_ref}*({base_wacc_cell_ref}/{wacc_cell})"
            )
            # Highlight base cross-point using index positions
            cell_fill = FILL_CROSS if wi == base_wacc_idx and ji == base_tgr_idx else None

            set_cell(ws, r, 3 + ji, formula,
                     font=_clone_font(FONT_FORMULA, fn),
                     fill=cell_fill,
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf)

    # --- Matrix 2: WACC x TGR -> Per Share Value ---
    r = m1_start_row + len(wacc_values) + 2
    set_cell(ws, r, 2, t("sens_wacc_tgr_ps", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)
    r += 1
    set_cell(ws, r, 2, f"{t('wacc', lang)} \\ {t('terminal_growth', lang)}",
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY,
             alignment=ALIGN_CENTER,
             border=BORDER_HEADER)
    for ji, tgr in enumerate(tgr_values):
        set_cell(ws, r, 3 + ji, tgr,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_DEEP_NAVY,
                 alignment=ALIGN_CENTER,
                 border=BORDER_HEADER,
                 number_format=NUM_PCT)

    m2_start_row = r + 1
    # Base WACC cell reference for matrix 2
    base_wacc_row_m2 = m2_start_row + base_wacc_idx
    base_wacc_cell_ref_m2 = f"${get_column_letter(2)}${base_wacc_row_m2}"

    for wi, wacc in enumerate(wacc_values):
        r = m2_start_row + wi
        set_cell(ws, r, 2, wacc,
                 font=_clone_font(FONT_INPUT, fn),
                 alignment=ALIGN_CENTER,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
        for ji, tgr in enumerate(tgr_values):
            wacc_cell = f"${get_column_letter(2)}${r}"
            tgr_cell = f"${get_column_letter(3 + ji)}${m2_start_row - 1}"
            formula = (
                f"=((({ufcf_ref}*(1+{tgr_cell})/({wacc_cell}-{tgr_cell})"
                f"+{ebitda_ref}*{base_exit_mult})"
                f"/2/(1+{wacc_cell})^{n_f}"
                f"+{sum_pv_ref}*({base_wacc_cell_ref_m2}/{wacc_cell})"
                f"-{net_debt_ref})/{shares}"
            )
            cell_fill = FILL_CROSS if wi == base_wacc_idx and ji == base_tgr_idx else None

            set_cell(ws, r, 3 + ji, formula,
                     font=_clone_font(FONT_FORMULA, fn),
                     fill=cell_fill,
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf_dec)

    # --- Matrix 3: WACC x Exit Multiple -> Enterprise Value ---
    r = m2_start_row + len(wacc_values) + 2
    set_cell(ws, r, 2, t("sens_wacc_mult_ev", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)
    r += 1
    set_cell(ws, r, 2, f"{t('wacc', lang)} \\ {t('exit_multiple', lang)}",
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY,
             alignment=ALIGN_CENTER,
             border=BORDER_HEADER)
    for ji, mult in enumerate(mult_values):
        set_cell(ws, r, 3 + ji, mult,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_DEEP_NAVY,
                 alignment=ALIGN_CENTER,
                 border=BORDER_HEADER,
                 number_format=NUM_MULT)

    m3_start_row = r + 1
    # Base WACC cell reference for matrix 3
    base_wacc_row_m3 = m3_start_row + base_wacc_idx
    base_wacc_cell_ref_m3 = f"${get_column_letter(2)}${base_wacc_row_m3}"

    for wi, wacc in enumerate(wacc_values):
        r = m3_start_row + wi
        set_cell(ws, r, 2, wacc,
                 font=_clone_font(FONT_INPUT, fn),
                 alignment=ALIGN_CENTER,
                 border=BORDER_THIN,
                 number_format=NUM_PCT)
        for ji, mult in enumerate(mult_values):
            wacc_cell = f"${get_column_letter(2)}${r}"
            mult_cell = f"${get_column_letter(3 + ji)}${m3_start_row - 1}"

            # EV = PV_GGM_TV + PV_Exit_TV(adjusted mult) + adjusted PV_FCF
            formula = (
                f"=({ufcf_ref}*(1+{base_tgr})/({wacc_cell}-{base_tgr})"
                f"+{ebitda_ref}*{mult_cell})"
                f"/2/(1+{wacc_cell})^{n_f}"
                f"+{sum_pv_ref}*({base_wacc_cell_ref_m3}/{wacc_cell})"
            )
            cell_fill = FILL_CROSS if wi == base_wacc_idx and ji == base_mult_idx else None

            set_cell(ws, r, 3 + ji, formula,
                     font=_clone_font(FONT_FORMULA, fn),
                     fill=cell_fill,
                     alignment=ALIGN_RIGHT,
                     border=BORDER_THIN,
                     number_format=cf)

    # --- Matrix 4: Scenario Comparison ---
    r = m3_start_row + len(wacc_values) + 2
    set_cell(ws, r, 2, t("scenario_comparison", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)
    r += 1
    comp_headers = [t("key_assumptions", lang),
                    t("scenario_bear", lang),
                    t("scenario_base", lang),
                    t("scenario_bull", lang)]
    write_header_row(ws, r, 2, comp_headers, lang)
    r += 1

    comp_items = [
        (t("revenue_cagr", lang), "cagr", NUM_PCT),
        (t("wacc", lang), "wacc", NUM_PCT2),
        (t("terminal_growth", lang), "tgr", NUM_PCT),
        (t("exit_multiple", lang), "exit_mult", NUM_MULT),
        (t("enterprise_value", lang), "ev", cf),
        (t("equity_value", lang), "equity", cf),
        (t("per_share_value", lang), "ps", cf_dec),
        (t("implied_ev_rev", lang), "ev_rev", NUM_MULT),
        (t("implied_ev_ebitda", lang), "ev_ebitda", NUM_MULT),
    ]

    for ci, (label, key, num_fmt) in enumerate(comp_items):
        set_cell(ws, r, 2, label,
                 font=_clone_font(FONT_LABEL, fn),
                 alignment=ALIGN_LEFT,
                 border=BORDER_THIN)
        for si, scenario in enumerate(scenarios_order):
            layout = dcf_layouts[scenario]
            cell_ref = layout.get(key)
            cell_fill = _scenario_fill(scenario)
            if cell_ref:
                formula = f"='{layout['sheet_name']}'!{cell_ref}"
                set_cell(ws, r, 3 + si, formula,
                         font=_clone_font(FONT_FORMULA, fn),
                         fill=cell_fill,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN,
                         number_format=num_fmt)
        r += 1

    return ws


# ============================================================
# 8. SOTP (Sum-of-the-Parts) VALUATION SHEET
# ============================================================

def write_sotp_sheet(wb: Workbook, data: dict, lang: str = "zh"):
    """Sheet: SOTP (Sum-of-the-Parts) Valuation.

    Data structure expected in data["sotp"]:
        segments: list of dicts with keys:
            name, revenue_ltm, ebitda_ltm, ev_multiple, ev, method, note
        net_debt: float
        minority_interest: float
        shares_outstanding: float (millions)
    """
    sotp_data = data.get("sotp")
    if not sotp_data:
        return None

    ws = wb.create_sheet(title="SOTP")
    ws.sheet_properties.tabColor = "BF8F00"  # Dark gold

    fn = _font_for_lang(lang)
    cur = data.get("currency", "USD")
    cf = currency_fmt(cur)
    cf_dec = currency_dec_fmt(cur)

    segments = sotp_data.get("segments", [])
    net_debt = sotp_data.get("net_debt", 0)
    minority_interest = sotp_data.get("minority_interest", 0)
    shares = sotp_data.get("shares_outstanding", data.get("shares_outstanding", 1000))

    set_col_widths(ws, {1: 2, 2: 22, 3: 16, 4: 16, 5: 16, 6: 12, 7: 18, 8: 22, 9: 3})

    # Title
    r = 1
    set_cell(ws, r, 2, t("sotp_title", lang),
             font=Font(name=fn, size=14, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)

    # Subtitle
    r = 2
    set_cell(ws, r, 2, f"{currency_symbol(cur)}M",
             font=_clone_font(FONT_NOTE, fn),
             alignment=ALIGN_LEFT)

    # Headers
    r = 4
    headers = [
        t("sotp_segment", lang),
        t("sotp_ltm_revenue", lang),
        t("sotp_ltm_ebitda", lang),
        t("sotp_method", lang),
        t("sotp_multiple", lang),
        t("sotp_ev", lang),
        t("sotp_note", lang),
    ]
    write_header_row(ws, r, 2, headers, lang)

    # Segment data rows
    data_start_row = 5
    for si, seg in enumerate(segments):
        r = data_start_row + si
        row_fill = _alt_row_fill(si)

        set_cell(ws, r, 2, seg.get("name", ""),
                 font=_clone_font(FONT_LABEL, fn),
                 fill=row_fill, alignment=ALIGN_LEFT, border=BORDER_THIN)

        set_cell(ws, r, 3, seg.get("revenue_ltm", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=cf)

        set_cell(ws, r, 4, seg.get("ebitda_ltm", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=cf)

        set_cell(ws, r, 5, seg.get("method", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_CENTER, border=BORDER_THIN)

        set_cell(ws, r, 6, seg.get("ev_multiple", ""),
                 font=_clone_font(FONT_INPUT, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_MULT2)

        # EV column: formula based on method
        method = seg.get("method", "")
        if "EBITDA" in method:
            ev_formula = f"=D{r}*F{r}"  # EBITDA x Multiple
        elif "Revenue" in method:
            ev_formula = f"=C{r}*F{r}"  # Revenue x Multiple
        else:
            ev_formula = f"=D{r}*F{r}"  # Default: EBITDA x Multiple
        set_cell(ws, r, 7, ev_formula,
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=cf)

        set_cell(ws, r, 8, seg.get("note", ""),
                 font=_clone_font(FONT_NOTE, fn),
                 fill=row_fill, alignment=ALIGN_LEFT, border=BORDER_THIN)

    # Summary section
    seg_end_row = data_start_row + len(segments) - 1
    r = seg_end_row + 2  # blank row gap

    # Total EV (formula sum)
    set_cell(ws, r, 2, t("sotp_total_ev", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
    ev_col_l = get_column_letter(7)
    set_cell(ws, r, 7, f"=SUM({ev_col_l}{data_start_row}:{ev_col_l}{seg_end_row})",
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
             number_format=cf)
    total_ev_row = r

    r += 1
    set_cell(ws, r, 2, t("sotp_less_net_debt", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
    set_cell(ws, r, 7, net_debt,
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
             number_format=cf)

    r += 1
    set_cell(ws, r, 2, t("sotp_less_minority", lang),
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
    set_cell(ws, r, 7, minority_interest,
             font=_clone_font(FONT_OUTPUT, fn),
             fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
             number_format=cf)

    r += 1
    set_cell(ws, r, 2, t("sotp_equity_value", lang),
             font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
             fill=FILL_GOLD, alignment=ALIGN_LEFT, border=BORDER_THIN)
    equity_row = r
    set_cell(ws, r, 7,
             f"={ev_col_l}{total_ev_row}-{ev_col_l}{total_ev_row+1}-{ev_col_l}{total_ev_row+2}",
             font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
             fill=FILL_GOLD, alignment=ALIGN_RIGHT, border=BORDER_THIN,
             number_format=cf)

    r += 1
    set_cell(ws, r, 2, t("sotp_shares", lang),
             font=_clone_font(FONT_LABEL, fn),
             fill=FILL_LIGHT_GRAY, alignment=ALIGN_LEFT, border=BORDER_THIN)
    set_cell(ws, r, 7, shares,
             font=_clone_font(FONT_INPUT, fn),
             fill=FILL_LIGHT_GRAY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
             number_format=NUM_CURRENCY)

    r += 1
    set_cell(ws, r, 2, t("sotp_per_share", lang),
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             fill=FILL_GOLD, alignment=ALIGN_LEFT, border=BORDER_THIN)
    set_cell(ws, r, 7,
             f"={ev_col_l}{equity_row}/{ev_col_l}{equity_row+1}",
             font=Font(name=fn, size=11, bold=True, color=DEEP_NAVY),
             fill=FILL_GOLD, alignment=ALIGN_RIGHT, border=BORDER_THIN,
             number_format=cf_dec)

    # Apply borders to summary columns
    apply_border_range(ws, total_ev_row, r, 2, 8, BORDER_THIN)

    return ws


# ============================================================
# 9. KEY OPERATIONAL KPIs SHEET
# ============================================================

def write_kpi_sheet(wb: Workbook, data: dict, lang: str = "zh"):
    """Sheet: Key Operational KPIs dashboard.

    Data structure expected in data["operational_kpis"]:
        metrics: list of dicts with keys:
            name, unit, historical (list), forecast (list)
    """
    kpi_data = data.get("operational_kpis")
    if not kpi_data:
        return None

    metrics = kpi_data.get("metrics", [])
    if not metrics:
        return None

    ws = wb.create_sheet(title=t("kpi_title", lang).split(" /")[0].strip() if lang == "zh" else "KPIs")
    ws.sheet_properties.tabColor = "4472C4"  # Blue

    fn = _font_for_lang(lang)

    # Determine column layout
    n_hist = max(len(m.get("historical", [])) for m in metrics)
    n_fcst = max(len(m.get("forecast", [])) for m in metrics)

    # Column layout: [spacer, Metric, Unit, Hist1..HistN, Fcst1..FcstN]
    col_start = 2
    metric_col = col_start
    unit_col = col_start + 1
    hist_start = col_start + 2
    fcst_start = hist_start + n_hist

    set_col_widths(ws, {1: 2, metric_col: 22, unit_col: 10})
    for i in range(n_hist + n_fcst):
        ws.column_dimensions[get_column_letter(hist_start + i)].width = 12

    # Title
    r = 1
    set_cell(ws, r, metric_col, t("kpi_title", lang),
             font=Font(name=fn, size=14, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)

    # Header row 1: merged labels for Historical / Forecast
    r = 3
    set_cell(ws, r, metric_col, t("kpi_metric", lang),
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY, alignment=ALIGN_CENTER, border=BORDER_HEADER)
    set_cell(ws, r, unit_col, t("kpi_unit", lang),
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY, alignment=ALIGN_CENTER, border=BORDER_HEADER)

    # Historical group header
    if n_hist > 0:
        set_cell(ws, r, hist_start, f"{t('kpi_hist_prefix', lang)} ({n_hist}Y)",
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_MED_BLUE, alignment=ALIGN_CENTER, border=BORDER_HEADER)
        ws.merge_cells(start_row=r, start_column=hist_start,
                        end_row=r, end_column=hist_start + n_hist - 1)

    # Forecast group header
    if n_fcst > 0:
        set_cell(ws, r, fcst_start, f"{t('kpi_fcst_prefix', lang)} ({n_fcst}Y)",
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_MED_BLUE, alignment=ALIGN_CENTER, border=BORDER_HEADER)
        ws.merge_cells(start_row=r, start_column=fcst_start,
                        end_row=r, end_column=fcst_start + n_fcst - 1)

    # Header row 2: year labels (placeholder Y-4, Y-3... or leave blank for user fill)
    r = 4
    set_cell(ws, r, metric_col, "",
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY, alignment=ALIGN_CENTER, border=BORDER_HEADER)
    set_cell(ws, r, unit_col, "",
             font=_clone_font(FONT_HEADER, fn),
             fill=FILL_DEEP_NAVY, alignment=ALIGN_CENTER, border=BORDER_HEADER)

    # Use historical years from data if available
    hist_years = data.get("historical", {}).get("years", [])
    fcst_years = data.get("forecast_years", [])

    for i in range(n_hist):
        col = hist_start + i
        label = hist_years[-n_hist + i] if len(hist_years) >= n_hist else f"H{i+1}"
        set_cell(ws, r, col, label,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_DEEP_NAVY, alignment=ALIGN_CENTER, border=BORDER_HEADER)

    for i in range(n_fcst):
        col = fcst_start + i
        label = fcst_years[i] if i < len(fcst_years) else f"F{i+1}"
        set_cell(ws, r, col, label,
                 font=_clone_font(FONT_HEADER, fn),
                 fill=FILL_DEEP_NAVY, alignment=ALIGN_CENTER, border=BORDER_HEADER)

    # Data rows
    data_start = 5
    for mi, metric in enumerate(metrics):
        r = data_start + mi
        row_fill = _alt_row_fill(mi)

        set_cell(ws, r, metric_col, metric.get("name", ""),
                 font=_clone_font(FONT_LABEL, fn),
                 fill=row_fill, alignment=ALIGN_LEFT, border=BORDER_THIN)

        set_cell(ws, r, unit_col, metric.get("unit", ""),
                 font=_clone_font(FONT_NOTE, fn),
                 fill=row_fill, alignment=ALIGN_CENTER, border=BORDER_THIN)

        # Historical values (green font for actuals)
        historical = metric.get("historical", [])
        for i, val in enumerate(historical):
            col = hist_start + i
            set_cell(ws, r, col, val,
                     font=_clone_font(FONT_ACTUAL, fn),
                     fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                     number_format=NUM_CURRENCY_DEC if isinstance(val, float) else NUM_CURRENCY)

        # Forecast values (blue font for estimates)
        forecast = metric.get("forecast", [])
        for i, val in enumerate(forecast):
            col = fcst_start + i
            set_cell(ws, r, col, val,
                     font=_clone_font(FONT_INPUT, fn),
                     fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                     number_format=NUM_CURRENCY_DEC if isinstance(val, float) else NUM_CURRENCY)

    # Add a separator line between historical and forecast
    sep_col = fcst_start - 1
    for mi in range(len(metrics)):
        r = data_start + mi
        cell = ws.cell(row=r, column=sep_col)
        cell.border = Border(
            left=Side(style="thin", color=MED_GRAY),
            right=Side(style="medium", color=DEEP_NAVY),
            top=Side(style="thin", color=MED_GRAY),
            bottom=Side(style="thin", color=MED_GRAY),
        )

    return ws


# ============================================================
# 10. HISTORICAL PE BAND SHEET
# ============================================================

def write_pe_band_sheet(wb: Workbook, data: dict, lang: str = "zh"):
    """Sheet: Historical PE Band analysis.

    Data structure expected in data["pe_band"]:
        current_pe: float
        pe_history: list of dicts with keys: period, high, low, close, eps
        median_pe: float
        forward_pe: float (optional)
    """
    pe_data = data.get("pe_band")
    if not pe_data:
        return None

    ws = wb.create_sheet(title="PE Band")
    ws.sheet_properties.tabColor = "ED7D31"  # Orange

    fn = _font_for_lang(lang)

    set_col_widths(ws, {1: 2, 2: 14, 3: 12, 4: 12, 5: 12, 6: 12, 7: 14, 8: 14, 9: 3})

    # Title
    r = 1
    set_cell(ws, r, 2, t("pe_band_title", lang),
             font=Font(name=fn, size=14, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)

    # Key metrics summary
    r = 3
    summary_items = [
        (t("pe_band_current", lang), pe_data.get("current_pe", ""), NUM_MULT),
        (t("pe_band_forward", lang), pe_data.get("forward_pe", ""), NUM_MULT),
        (t("pe_band_median", lang), pe_data.get("median_pe", ""), NUM_MULT),
    ]

    for label, val, fmt in summary_items:
        if val != "":
            set_cell(ws, r, 2, label,
                     font=_clone_font(FONT_LABEL, fn),
                     fill=FILL_LIGHT_GRAY, alignment=ALIGN_LEFT, border=BORDER_THIN)
            set_cell(ws, r, 3, val,
                     font=_clone_font(FONT_OUTPUT, fn),
                     fill=FILL_LIGHT_GRAY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                     number_format=fmt)
            r += 1

    # PE History table
    r += 1
    headers = [
        t("pe_band_period", lang),
        t("pe_band_high", lang),
        t("pe_band_low", lang),
        t("pe_band_close", lang),
        t("pe_band_eps", lang),
        "Close PE",
    ]
    write_header_row(ws, r, 2, headers, lang)
    header_row = r
    r += 1

    pe_history = pe_data.get("pe_history", [])
    data_start = r
    for pi, period in enumerate(pe_history):
        row_fill = _alt_row_fill(pi)

        set_cell(ws, r, 2, period.get("period", ""),
                 font=_clone_font(FONT_LABEL, fn),
                 fill=row_fill, alignment=ALIGN_CENTER, border=BORDER_THIN)

        set_cell(ws, r, 3, period.get("high", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_CURRENCY)

        set_cell(ws, r, 4, period.get("low", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_CURRENCY)

        set_cell(ws, r, 5, period.get("close", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_CURRENCY)

        set_cell(ws, r, 6, period.get("eps", ""),
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_CURRENCY_DEC)

        # Close PE = Close / EPS (formula)
        close_col = get_column_letter(5)
        eps_col = get_column_letter(6)
        set_cell(ws, r, 7, f'=IF({eps_col}{r}=0,"",{close_col}{r}/{eps_col}{r})',
                 font=_clone_font(FONT_FORMULA, fn),
                 fill=row_fill, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_MULT)

        r += 1

    data_end = r - 1

    # Summary statistics for PE history
    if pe_history:
        r += 1
        set_cell(ws, r, 2, t("comp_median", lang),
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
        pe_col = get_column_letter(7)
        set_cell(ws, r, 7, f"=MEDIAN({pe_col}{data_start}:{pe_col}{data_end})",
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_MULT)

        r += 1
        set_cell(ws, r, 2, t("comp_mean", lang),
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
        set_cell(ws, r, 7, f"=AVERAGE({pe_col}{data_start}:{pe_col}{data_end})",
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_MULT)

        r += 1
        set_cell(ws, r, 2, "Max",
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
        set_cell(ws, r, 7, f"=MAX({pe_col}{data_start}:{pe_col}{data_end})",
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_MULT)

        r += 1
        set_cell(ws, r, 2, "Min",
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_LEFT, border=BORDER_THIN)
        set_cell(ws, r, 7, f"=MIN({pe_col}{data_start}:{pe_col}{data_end})",
                 font=_clone_font(FONT_OUTPUT, fn),
                 fill=FILL_SUMMARY, alignment=ALIGN_RIGHT, border=BORDER_THIN,
                 number_format=NUM_MULT)

    # Note
    r += 2
    set_cell(ws, r, 2, t("pe_band_note", lang),
             font=_clone_font(FONT_SMALL, fn),
             alignment=ALIGN_LEFT)

    return ws


# ============================================================
# 11. COMPARABLE COMPANIES SHEET (Enhanced with multi-market grouping)
# ============================================================

def write_comps_sheet(wb: Workbook, data: dict, lang: str = "zh"):
    """Sheet: Comparable Companies analysis.

    Enhanced V7: supports multi-market grouping via "market" field.
    If any company has a "market" field (e.g. "US", "HK", "A"),
    companies are grouped by market with per-group median rows.
    Otherwise, falls back to the original flat layout.
    """
    ws = wb.create_sheet(title=t("comps", lang))
    ws.sheet_properties.tabColor = "70AD47"

    fn = _font_for_lang(lang)
    cur = data.get("currency", "USD")
    cf = currency_fmt(cur)

    comps = data.get("comparable_companies", [])
    n_comps = len(comps)

    # Check if multi-market grouping is needed
    has_market = any(c.get("market") for c in comps)

    # Market label mapping
    market_labels = {
        "US": t("comp_market_us", lang),
        "HK": t("comp_market_hk", lang),
        "A": t("comp_market_a", lang),
    }

    col_keys = [
        ("comp_name", 24),
        ("comp_ticker", 10),
        ("comp_mcap", 14),
        ("comp_ev", 14),
        ("comp_ev_rev_ltm", 14),
        ("comp_ev_rev_ntm", 14),
        ("comp_ev_ebitda_ltm", 14),
        ("comp_ev_ebitda_ntm", 14),
        ("comp_cagr", 12),
        ("comp_note", 20),
    ]

    set_col_widths(ws, {1: 2})
    for i, (_, w) in enumerate(col_keys):
        ws.column_dimensions[get_column_letter(2 + i)].width = w

    # Title
    r = 1
    set_cell(ws, r, 2, t("comps", lang),
             font=Font(name=fn, size=14, bold=True, color=DEEP_NAVY),
             alignment=ALIGN_LEFT)

    data_keys = ["name", "ticker", "mcap", "ev",
                 "ev_rev_ltm", "ev_rev_ntm", "ev_ebitda_ltm", "ev_ebitda_ntm",
                 "cagr_3y", "note"]

    if not has_market:
        # Original flat layout (backward compatible)
        r = 2
        headers = [t(k, lang) for k, _ in col_keys]
        write_header_row(ws, r, 2, headers, lang)

        for ci, comp in enumerate(comps):
            r = 3 + ci
            row_fill = _alt_row_fill(ci)
            for di, dk in enumerate(data_keys):
                val = comp.get(dk, "")
                col = 2 + di
                num_fmt = None
                font = _clone_font(FONT_FORMULA, fn)

                if dk in ("mcap", "ev"):
                    num_fmt = cf
                elif dk in ("ev_rev_ltm", "ev_rev_ntm", "ev_ebitda_ltm", "ev_ebitda_ntm"):
                    num_fmt = NUM_MULT
                elif dk == "cagr_3y":
                    num_fmt = NUM_PCT

                set_cell(ws, r, col, val,
                         font=font,
                         fill=row_fill,
                         alignment=ALIGN_RIGHT if num_fmt else ALIGN_LEFT,
                         border=BORDER_THIN,
                         number_format=num_fmt)

        # Median and Mean rows
        if n_comps > 0:
            median_row = 3 + n_comps
            mean_row = median_row + 1

            set_cell(ws, median_row, 2, t("comp_median", lang),
                     font=_clone_font(FONT_OUTPUT, fn),
                     fill=FILL_SUMMARY,
                     alignment=ALIGN_LEFT,
                     border=BORDER_THIN)
            set_cell(ws, mean_row, 2, t("comp_mean", lang),
                     font=_clone_font(FONT_OUTPUT, fn),
                     fill=FILL_SUMMARY,
                     alignment=ALIGN_LEFT,
                     border=BORDER_THIN)

            for di in range(2, 2 + len(data_keys)):
                col_l = get_column_letter(di)
                first_data = f"{col_l}3"
                last_data = f"{col_l}{2 + n_comps}"

                median_formula = f"=MEDIAN({first_data}:{last_data})"
                mean_formula = f"=AVERAGE({first_data}:{last_data})"

                set_cell(ws, median_row, di, median_formula,
                         font=_clone_font(FONT_OUTPUT, fn),
                         fill=FILL_SUMMARY,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN)
                set_cell(ws, mean_row, di, mean_formula,
                         font=_clone_font(FONT_OUTPUT, fn),
                         fill=FILL_SUMMARY,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN)
    else:
        # Multi-market grouped layout
        # Group companies by market
        market_order = ["US", "HK", "A"]
        grouped = {}
        ungrouped = []
        for comp in comps:
            mkt = comp.get("market", "")
            if mkt in market_order:
                grouped.setdefault(mkt, []).append(comp)
            else:
                ungrouped.append(comp)

        r = 2
        for mkt in market_order:
            mkt_comps = grouped.get(mkt, [])
            if not mkt_comps:
                continue

            # Market group header
            mkt_label = market_labels.get(mkt, mkt)
            set_cell(ws, r, 2, mkt_label,
                     font=Font(name=fn, size=11, bold=True, color=WHITE),
                     fill=FILL_MED_BLUE,
                     alignment=ALIGN_LEFT,
                     border=BORDER_HEADER)
            # Merge across all columns
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=2 + len(col_keys) - 1)
            r += 1

            # Column headers for this group
            headers = [t(k, lang) for k, _ in col_keys]
            write_header_row(ws, r, 2, headers, lang)
            r += 1

            group_start = r
            for ci, comp in enumerate(mkt_comps):
                row_fill = _alt_row_fill(ci)
                for di, dk in enumerate(data_keys):
                    val = comp.get(dk, "")
                    col = 2 + di
                    num_fmt = None
                    font = _clone_font(FONT_FORMULA, fn)

                    if dk in ("mcap", "ev"):
                        num_fmt = cf
                    elif dk in ("ev_rev_ltm", "ev_rev_ntm", "ev_ebitda_ltm", "ev_ebitda_ntm"):
                        num_fmt = NUM_MULT
                    elif dk == "cagr_3y":
                        num_fmt = NUM_PCT

                    set_cell(ws, r, col, val,
                             font=font,
                             fill=row_fill,
                             alignment=ALIGN_RIGHT if num_fmt else ALIGN_LEFT,
                             border=BORDER_THIN,
                             number_format=num_fmt)
                r += 1
            group_end = r - 1

            # Sub-median row for this market
            set_cell(ws, r, 2, f"  {t('comp_sub_median', lang)} ({mkt_label})",
                     font=_clone_font(FONT_OUTPUT, fn),
                     fill=FILL_SUMMARY,
                     alignment=ALIGN_LEFT,
                     border=BORDER_THIN)
            for di in range(3, 2 + len(data_keys)):
                col_l = get_column_letter(di)
                set_cell(ws, r, di, f"=MEDIAN({col_l}{group_start}:{col_l}{group_end})",
                         font=_clone_font(FONT_OUTPUT, fn),
                         fill=FILL_SUMMARY,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN)
            r += 1
            r += 1  # blank row between groups

        # Ungrouped companies (if any)
        if ungrouped:
            set_cell(ws, r, 2, t("comp_market_other", lang),
                     font=Font(name=fn, size=11, bold=True, color=WHITE),
                     fill=FILL_MED_BLUE,
                     alignment=ALIGN_LEFT,
                     border=BORDER_HEADER)
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=2 + len(col_keys) - 1)
            r += 1

            headers = [t(k, lang) for k, _ in col_keys]
            write_header_row(ws, r, 2, headers, lang)
            r += 1

            group_start = r
            for ci, comp in enumerate(ungrouped):
                row_fill = _alt_row_fill(ci)
                for di, dk in enumerate(data_keys):
                    val = comp.get(dk, "")
                    col = 2 + di
                    num_fmt = None
                    font = _clone_font(FONT_FORMULA, fn)

                    if dk in ("mcap", "ev"):
                        num_fmt = cf
                    elif dk in ("ev_rev_ltm", "ev_rev_ntm", "ev_ebitda_ltm", "ev_ebitda_ntm"):
                        num_fmt = NUM_MULT
                    elif dk == "cagr_3y":
                        num_fmt = NUM_PCT

                    set_cell(ws, r, col, val,
                             font=font,
                             fill=row_fill,
                             alignment=ALIGN_RIGHT if num_fmt else ALIGN_LEFT,
                             border=BORDER_THIN,
                             number_format=num_fmt)
                r += 1
            group_end = r - 1

            set_cell(ws, r, 2, f"  {t('comp_sub_median', lang)}",
                     font=_clone_font(FONT_OUTPUT, fn),
                     fill=FILL_SUMMARY,
                     alignment=ALIGN_LEFT,
                     border=BORDER_THIN)
            for di in range(3, 2 + len(data_keys)):
                col_l = get_column_letter(di)
                set_cell(ws, r, di, f"=MEDIAN({col_l}{group_start}:{col_l}{group_end})",
                         font=_clone_font(FONT_OUTPUT, fn),
                         fill=FILL_SUMMARY,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN)
            r += 1
            r += 1

        # Overall median row (across all markets)
        set_cell(ws, r, 2, t("comp_median", lang),
                 font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
                 fill=FILL_GOLD,
                 alignment=ALIGN_LEFT,
                 border=BORDER_THIN)
        # For overall median, we reference all data rows (skip group headers and median rows)
        # Collect all data row numbers
        all_data_rows = []
        tmp_r = 3  # data starts after title and first header
        for mkt in market_order:
            mkt_comps = grouped.get(mkt, [])
            if not mkt_comps:
                continue
            tmp_r += 2  # group header + column header
            for ci in range(len(mkt_comps)):
                all_data_rows.append(tmp_r)
                tmp_r += 1
            tmp_r += 2  # sub-median + blank
        if ungrouped:
            tmp_r += 2
            for ci in range(len(ungrouped)):
                all_data_rows.append(tmp_r)
                tmp_r += 1

        if all_data_rows:
            for di in range(3, 2 + len(data_keys)):
                col_l = get_column_letter(di)
                refs = ",".join(f"{col_l}{row_n}" for row_n in all_data_rows)
                set_cell(ws, r, di, f"=MEDIAN({refs})",
                         font=Font(name=fn, size=10, bold=True, color=DEEP_NAVY),
                         fill=FILL_GOLD,
                         alignment=ALIGN_RIGHT,
                         border=BORDER_THIN)

    return ws


# ============================================================
# 12. CHART HELPER FUNCTIONS (openpyxl native charts)
# ============================================================

def _chart_title_rich_text(text: str, font_size: int = CHART_TITLE_SIZE):
    """Build a RichText object for chart title."""
    return RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=font_size, b=True, solidFill=DEEP_NAVY)
            ),
            endParaRPr=CharacterProperties(sz=font_size, b=True, solidFill=DEEP_NAVY),
        )]
    )


def _apply_chart_common(chart, title_text: str):
    """Apply common MS-style formatting to a chart object."""
    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'b'
    chart.legend.includeInLayout = False
    chart.style = 2  # clean white background style

    # Title
    chart.title = title_text
    chart.title.txPr = _chart_title_rich_text(title_text)

    # Legend font size
    if chart.legend.txPr is None:
        chart.legend.txPr = RichText(
            p=[ChartParagraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
                ),
            )]
        )

    # Y-axis / X-axis styling
    for axis in [chart.y_axis, chart.x_axis]:
        axis.tickLblPos = 'low'
        axis.delete = False
        if hasattr(axis, 'majorGridlines') and axis.majorGridlines:
            axis.majorGridlines.spPr = None  # reset
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.line import LineProperties, NoFill
            axis.majorGridlines.spPr = GraphicalProperties()


def _add_revenue_ebitda_chart(wb, sheet_name, theme="classic"):
    """Add Revenue & EBITDA dual bar chart to the specified DCF sheet.

    Reads Total Revenue row and EBITDA row from the sheet data area.
    Places chart to the right of the data columns.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # Determine data range from the sheet
    # Row 2 = year headers, col_start = 3 (C)
    # We need to find total_rev_row and ebitda_row by scanning column B
    total_rev_row = None
    ebitda_row = None
    year_header_row = 2
    col_start = 3

    def _row_has_numeric_data(ws, row, col_start, col_end):
        """Check if a row has at least one non-None value (numeric or formula) in the given column range."""
        for c in range(col_start, col_end + 1):
            val = ws.cell(row=row, column=c).value
            if val is not None:
                return True
        return False

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and isinstance(cell_val, str):
            if "Total Revenue" in cell_val or "总收入" in cell_val:
                # Only accept rows that have actual data (numeric or formula)
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    total_rev_row = row
            elif cell_val.strip() == "EBITDA" or cell_val.strip() == "息税折旧摊销前利润":
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    ebitda_row = row

    if total_rev_row is None or ebitda_row is None:
        return

    # Find last data column
    last_col = col_start
    for c in range(col_start, ws.max_column + 1):
        if ws.cell(row=year_header_row, column=c).value is not None:
            last_col = c

    n_cols = last_col - col_start + 1

    # Categories (year headers from row 2)
    cats = Reference(ws, min_col=col_start, max_col=last_col,
                     min_row=year_header_row, max_row=year_header_row)

    # Revenue series
    rev_data = Reference(ws, min_col=col_start, max_col=last_col,
                          min_row=total_rev_row, max_row=total_rev_row)

    # EBITDA series
    ebitda_data = Reference(ws, min_col=col_start, max_col=last_col,
                            min_row=ebitda_row, max_row=ebitda_row)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Revenue & EBITDA ($M)"
    chart.y_axis.title = "$M"
    chart.x_axis.title = None
    chart.style = 2

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'b'
    chart.legend.includeInLayout = False

    # Title formatting
    chart.title = "Revenue & EBITDA ($M)"
    chart.title.txPr = _chart_title_rich_text("Revenue & EBITDA ($M)")

    # Legend font
    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    # Gap width
    chart.gapWidth = CHART_GAP_WIDTH

    # Add series with colors
    chart.add_data(rev_data, titles_from_data=False, from_rows=True)
    chart.add_data(ebitda_data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    # Series 0 = Revenue
    s0 = chart.series[0]
    s0.title = SeriesLabel(v="Revenue ($M)")
    s0.graphicalProperties.solidFill = "1F3864"

    # Series 1 = EBITDA
    s1 = chart.series[1]
    s1.title = SeriesLabel(v="EBITDA ($M)")
    s1.graphicalProperties.solidFill = "2E75B6"

    # Data labels
    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = '#,##0'
        s.dLbls.txPr = RichText(
            p=[ChartParagraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
                ),
            )]
        )

    # Y-axis formatting
    chart.y_axis.numFmt = '#,##0'
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    # Grid lines - light gray
    chart.y_axis.majorGridlines = None  # clean default

    # Place chart to the right of data
    anchor_col = last_col + 2
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}1")


def _add_margin_line_chart(wb, sheet_name, theme="classic"):
    """Add margin line chart (Gross Margin, EBITDA Margin, Net Margin) to DCF sheet.

    Reads margin rows from the sheet data area.
    Places chart below the revenue/EBITDA bar chart.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # Find margin rows by scanning column B
    ebitda_margin_row = None
    year_header_row = 2
    col_start = 3

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and isinstance(cell_val, str):
            if "EBITDA Margin" in cell_val or "EBITDA利润率" in cell_val:
                ebitda_margin_row = row
                break

    if ebitda_margin_row is None:
        return

    # Find last data column
    last_col = col_start
    for c in range(col_start, ws.max_column + 1):
        if ws.cell(row=year_header_row, column=c).value is not None:
            last_col = c

    # We need Gross Margin and Net Margin rows too.
    # Gross Margin is not always present; we'll compute from data if possible.
    # For now, use EBITDA Margin and add placeholder rows for Gross/Net.
    # Strategy: write helper data rows in a hidden area, or use existing rows.
    # We'll look for common patterns.

    gross_margin_row = None
    net_margin_row = None

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and isinstance(cell_val, str):
            cv = cell_val.strip()
            if cv in ("Gross Margin", "Gross Profit Margin", "毛利率", "毛利润率"):
                gross_margin_row = row
            elif cv in ("Net Margin", "Net Profit Margin", "净利率", "净利润率"):
                net_margin_row = row

    # If gross/net margin rows don't exist, create them in a helper area
    # We'll write them to columns far right (col AA = 27+) as helper data
    helper_col_start = 27  # Column AA
    helper_row_base = max(ws.max_row + 2, 60)

    # Find EBIT row for gross margin calc, and NOPAT row for net margin
    ebit_row = None
    ebitda_row = None
    nopat_row = None
    total_rev_row = None

    def _row_has_numeric_data(ws, row, col_start, col_end):
        """Check if a row has at least one non-None value (numeric or formula) in the given column range."""
        for c in range(col_start, col_end + 1):
            val = ws.cell(row=row, column=c).value
            if val is not None:
                return True
        return False

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and isinstance(cell_val, str):
            cv = cell_val.strip()
            if cv in ("EBIT", "息税前利润"):
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    ebit_row = row
            elif cv == "EBITDA" or cv == "息税折旧摊销前利润":
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    ebitda_row = row
            elif cv in ("NOPAT", "税后营业利润"):
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    nopat_row = row
            elif "Total Revenue" in cv or "总收入" in cv:
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    total_rev_row = row

    # Write Gross Margin helper data
    gm_helper_row = helper_row_base
    ws.cell(row=gm_helper_row, column=helper_col_start, value="Gross Margin (helper)")
    if ebit_row and total_rev_row:
        for ci in range(col_start, last_col + 1):
            cl = get_column_letter(ci)
            hc = get_column_letter(helper_col_start + ci - col_start + 1)
            # Gross Margin approx = EBIT / Revenue (since we don't have COGS)
            # Use EBIT as proxy for gross profit
            formula = f"=IF({cl}{total_rev_row}=0,0,{cl}{ebit_row}/{cl}{total_rev_row})"
            c = ws.cell(row=gm_helper_row, column=helper_col_start + ci - col_start + 1,
                    value=formula)
            c.number_format = '0.0%'

    # Write Net Margin helper data
    nm_helper_row = helper_row_base + 1
    ws.cell(row=nm_helper_row, column=helper_col_start, value="Net Margin (helper)")
    if nopat_row and total_rev_row:
        for ci in range(col_start, last_col + 1):
            cl = get_column_letter(ci)
            hc = get_column_letter(helper_col_start + ci - col_start + 1)
            formula = f"=IF({cl}{total_rev_row}=0,0,{cl}{nopat_row}/{cl}{total_rev_row})"
            c = ws.cell(row=nm_helper_row, column=helper_col_start + ci - col_start + 1,
                    value=formula)
            c.number_format = '0.0%'

    # Write EBITDA Margin helper data (copy from existing)
    em_helper_row = helper_row_base + 2
    ws.cell(row=em_helper_row, column=helper_col_start, value="EBITDA Margin (helper)")
    if ebitda_margin_row and total_rev_row and ebitda_row:
        for ci in range(col_start, last_col + 1):
            cl = get_column_letter(ci)
            hc = get_column_letter(helper_col_start + ci - col_start + 1)
            formula = f"=IF({cl}{total_rev_row}=0,0,{cl}{ebitda_row}/{cl}{total_rev_row})"
            c = ws.cell(row=em_helper_row, column=helper_col_start + ci - col_start + 1,
                    value=formula)
            c.number_format = '0.0%'

    # Build chart from helper data
    n_data_cols = last_col - col_start + 1
    helper_last_col = helper_col_start + n_data_cols

    # Categories (year headers copied to helper area)
    for ci in range(n_data_cols):
        src_val = ws.cell(row=year_header_row, column=col_start + ci).value
        ws.cell(row=helper_row_base - 1, column=helper_col_start + 1 + ci, value=src_val)

    cats = Reference(ws, min_col=helper_col_start + 1, max_col=helper_last_col,
                     min_row=helper_row_base - 1, max_row=helper_row_base - 1)

    chart = LineChart()
    chart.title = "Profitability Margins"
    chart.style = 2

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'b'
    chart.legend.includeInLayout = False

    chart.title = "Profitability Margins"
    chart.title.txPr = _chart_title_rich_text("Profitability Margins")

    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    # Add three series: Gross Margin, EBITDA Margin, Net Margin
    margin_colors = ['1F3864', 'C8A951', '00AF50']
    margin_labels = ['Gross Margin', 'EBITDA Margin', 'Net Margin']
    margin_rows = [gm_helper_row, em_helper_row, nm_helper_row]

    for i, (mrow, mlabel, mcolor) in enumerate(zip(margin_rows, margin_labels, margin_colors)):
        data_ref = Reference(ws, min_col=helper_col_start + 1, max_col=helper_last_col,
                             min_row=mrow, max_row=mrow)
        chart.add_data(data_ref, titles_from_data=False, from_rows=True)
        s = chart.series[i]
        s.title = SeriesLabel(v=mlabel)
        s.graphicalProperties.line.solidFill = mcolor
        s.graphicalProperties.line.width = CHART_LINE_WIDTH
        # Marker
        s.marker.symbol = "circle"
        s.marker.size = 5
        s.marker.graphicalProperties.solidFill = mcolor
        # Data labels
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = '0.0%'
        s.dLbls.txPr = RichText(
            p=[ChartParagraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
                ),
            )]
        )

    chart.set_categories(cats)

    # Y-axis formatting
    chart.y_axis.numFmt = '0%'
    chart.y_axis.scaling.min = 0
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    # Place chart below the bar chart (row 18 or so, same anchor column)
    anchor_col = last_col + 2
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}18")


def _add_sotp_pie_chart(wb, sheet_name, theme="classic"):
    """Add SOTP pie chart showing segment enterprise values.

    Reads segment names (col B) and EV values (col G) from the SOTP sheet.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # SOTP layout: headers at row 4, data starts at row 5
    # Col B = segment name, Col G = EV
    header_row = 4
    data_start_row = 5

    # Find how many segments
    last_data_row = data_start_row
    for r in range(data_start_row, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is not None and \
           ws.cell(row=r, column=7).value is not None:
            last_data_row = r
        else:
            break

    n_segs = last_data_row - data_start_row + 1
    if n_segs <= 0:
        return

    # Categories (segment names)
    cats = Reference(ws, min_col=2, min_row=data_start_row, max_row=last_data_row)

    # Data (EV values)
    data = Reference(ws, min_col=7, min_row=data_start_row, max_row=last_data_row)

    chart = PieChart()
    chart.title = "SOTP - Enterprise Value by Segment"
    chart.style = 2

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'r'
    chart.legend.includeInLayout = False

    chart.title = "SOTP - Enterprise Value by Segment"
    chart.title.txPr = _chart_title_rich_text("SOTP - Enterprise Value by Segment")

    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    # Apply MS 8-color palette to slices
    series = chart.series[0]
    for i in range(n_segs):
        pt = DataPoint(idx=i)
        color_idx = i % len(CHART_PALETTE)
        pt.graphicalProperties.solidFill = CHART_PALETTE[color_idx]
        series.data_points.append(pt)

    # Data labels: show percentage
    series.dLbls = DataLabelList()
    series.dLbls.showPercent = True
    series.dLbls.showCatName = False
    series.dLbls.showVal = False
    series.dLbls.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
            ),
        )]
    )

    # Place chart to the right of data
    ws.add_chart(chart, "J4")


def _add_pe_band_chart(wb, sheet_name, theme="classic"):
    """Add PE Band area chart to the PE Band sheet.

    Reads PE High, PE Low, and Close PE columns from the PE Band sheet.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # PE Band layout: headers at the row with "Period/High/Low/Close/EPS/Close PE"
    # Find header row
    header_row = None
    for r in range(1, min(ws.max_row + 1, 20)):
        val = ws.cell(row=r, column=2).value
        if val and ("Period" in str(val) or "期间" in str(val)):
            header_row = r
            break

    if header_row is None:
        return

    data_start = header_row + 1

    # Find last data row
    last_data_row = data_start
    for r in range(data_start, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is not None:
            last_data_row = r
        else:
            break

    n_rows = last_data_row - data_start + 1
    if n_rows <= 0:
        return

    # Col B = Period, Col C = High, Col D = Low, Col G = Close PE
    # Categories (Period)
    cats = Reference(ws, min_col=2, min_row=data_start, max_row=last_data_row)

    # High band
    high_data = Reference(ws, min_col=3, min_row=data_start, max_row=last_data_row)

    # Low band
    low_data = Reference(ws, min_col=4, min_row=data_start, max_row=last_data_row)

    # Close PE
    close_pe_data = Reference(ws, min_col=7, min_row=data_start, max_row=last_data_row)

    # Create AreaChart for the band
    chart = AreaChart()
    chart.title = "Historical PE Band"
    chart.style = 2
    chart.grouping = "standard"

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'b'
    chart.legend.includeInLayout = False

    chart.title = "Historical PE Band"
    chart.title.txPr = _chart_title_rich_text("Historical PE Band")

    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    # Add series: High Band (area), Low Band (area), Close PE (line)
    chart.add_data(high_data, titles_from_data=False)
    chart.add_data(low_data, titles_from_data=False)
    chart.set_categories(cats)

    # High band series
    s_high = chart.series[0]
    s_high.title = SeriesLabel(v="PE High Band")
    s_high.graphicalProperties.solidFill = "2E75B6"
    s_high.graphicalProperties.line.solidFill = "2E75B6"

    # Low band series
    s_low = chart.series[1]
    s_low.title = SeriesLabel(v="PE Low Band")
    s_low.graphicalProperties.solidFill = "BDD7EE"
    s_low.graphicalProperties.line.solidFill = "BDD7EE"

    # Add Close PE as a line overlay
    line_chart = LineChart()
    line_chart.add_data(close_pe_data, titles_from_data=False)
    line_chart.set_categories(cats)

    s_close = line_chart.series[0]
    s_close.title = SeriesLabel(v="PE Close")
    s_close.graphicalProperties.line.solidFill = "1F3864"
    s_close.graphicalProperties.line.width = CHART_LINE_WIDTH
    s_close.marker.symbol = "diamond"
    s_close.marker.size = 5
    s_close.marker.graphicalProperties.solidFill = "1F3864"

    # Data labels for close PE
    s_close.dLbls = DataLabelList()
    s_close.dLbls.showVal = True
    s_close.dLbls.numFmt = '0.0"x"'
    s_close.dLbls.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
            ),
        )]
    )

    # Overlay line on area
    chart += line_chart

    # Y-axis formatting
    chart.y_axis.numFmt = '0.0"x"'
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    # Place chart to the right of data
    ws.add_chart(chart, "J3")


def _add_stacked_revenue_chart(wb, sheet_name, theme="classic"):
    """Add stacked bar chart showing revenue breakdown by business segment.

    Scans the DCF sheet for segment revenue rows (rows 4..4+n_seg-1 by convention,
    where each row in col B starts with two spaces indicating a segment name).
    Uses from_rows=True for openpyxl row-oriented data.
    Each segment gets a distinct color from CHART_PALETTE.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    year_header_row = 2
    col_start = 3

    # Find last data column from year headers
    last_col = col_start
    for c in range(col_start, ws.max_column + 1):
        if ws.cell(row=year_header_row, column=c).value is not None:
            last_col = c

    # Scan for segment revenue rows: rows where col B starts with "  " (two spaces)
    # and have numeric data, stopping before "Total Revenue"
    seg_rows = []
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and isinstance(cell_val, str):
            cv = cell_val.strip()
            # Stop at Total Revenue row
            if "Total Revenue" in cv or "总收入" in cv:
                break
            # Segment rows are indented with two spaces
            if cell_val.startswith("  ") and cv != "":
                # Verify row has numeric data
                has_data = False
                for c in range(col_start, last_col + 1):
                    val = ws.cell(row=row, column=c).value
                    if val is not None:
                        has_data = True
                        break
                if has_data:
                    seg_rows.append(row)

    if len(seg_rows) < 1:
        return

    # Categories (year headers)
    cats = Reference(ws, min_col=col_start, max_col=last_col,
                     min_row=year_header_row, max_row=year_header_row)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Revenue Breakdown by Segment ($M)"
    chart.style = 2

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'b'
    chart.legend.includeInLayout = False

    chart.title = "Revenue Breakdown by Segment ($M)"
    chart.title.txPr = _chart_title_rich_text("Revenue Breakdown by Segment ($M)")

    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    chart.gapWidth = CHART_GAP_WIDTH

    # Add each segment as a series using from_rows=True
    for i, seg_row in enumerate(seg_rows):
        data_ref = Reference(ws, min_col=col_start, max_col=last_col,
                             min_row=seg_row, max_row=seg_row)
        chart.add_data(data_ref, titles_from_data=False, from_rows=True)

        s = chart.series[i]
        # Use segment name (trimmed) as series title
        seg_name = ws.cell(row=seg_row, column=2).value.strip()
        s.title = SeriesLabel(v=seg_name)
        # Assign color from CHART_PALETTE
        color_idx = i % len(CHART_PALETTE)
        s.graphicalProperties.solidFill = CHART_PALETTE[color_idx]

    chart.set_categories(cats)

    # Y-axis formatting
    chart.y_axis.title = "$M"
    chart.y_axis.numFmt = '#,##0'
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    # Place chart below the margin line chart (row 35)
    anchor_col = last_col + 2
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}35")


def _add_donut_chart(wb, sheet_name, theme="classic"):
    """Add donut/ring chart for SOTP or market composition.

    Uses DoughnutChart from openpyxl if available, otherwise falls back to
    PieChart. Shows total enterprise value in center via a data label on the
    last data point.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # SOTP layout: headers at row 4, data starts at row 5
    # Col B = segment name, Col G = EV
    header_row = 4
    data_start_row = 5

    # Find how many segments
    last_data_row = data_start_row
    for r in range(data_start_row, ws.max_row + 1):
        if ws.cell(row=r, column=2).value is not None and \
           ws.cell(row=r, column=7).value is not None:
            last_data_row = r
        else:
            break

    n_segs = last_data_row - data_start_row + 1
    if n_segs <= 0:
        return

    # Categories (segment names)
    cats = Reference(ws, min_col=2, min_row=data_start_row, max_row=last_data_row)

    # Data (EV values)
    data = Reference(ws, min_col=7, min_row=data_start_row, max_row=last_data_row)

    # Try to use DoughnutChart (available in openpyxl >= 3.0)
    try:
        from openpyxl.chart import DoughnutChart
        chart = DoughnutChart()
        chart_type = "doughnut"
    except ImportError:
        chart = PieChart()
        chart_type = "pie"

    chart.title = "SOTP - Enterprise Value by Segment (Donut)"
    chart.style = 2

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'r'
    chart.legend.includeInLayout = False

    chart.title = "SOTP - Enterprise Value by Segment (Donut)"
    chart.title.txPr = _chart_title_rich_text("SOTP - Enterprise Value by Segment (Donut)")

    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    # Apply MS 8-color palette to slices
    series = chart.series[0]
    for i in range(n_segs):
        pt = DataPoint(idx=i)
        color_idx = i % len(CHART_PALETTE)
        pt.graphicalProperties.solidFill = CHART_PALETTE[color_idx]
        series.data_points.append(pt)

    # Data labels: show percentage
    series.dLbls = DataLabelList()
    series.dLbls.showPercent = True
    series.dLbls.showCatName = False
    series.dLbls.showVal = False
    series.dLbls.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
            ),
        )]
    )

    # For DoughnutChart, set hole size if supported
    if chart_type == "doughnut" and hasattr(chart, 'holeSize'):
        chart.holeSize = 50

    # Place chart below the existing pie chart
    ws.add_chart(chart, "J20")


def _add_grouped_scenario_chart(wb, sheet_name, theme="classic"):
    """Add grouped/clustered bar chart comparing Bear/Base/Bull scenarios.

    Reads key valuation metrics (Revenue, EBITDA, Equity Value, Price per Share)
    from each of the three DCF scenario sheets and displays them side by side.
    Each scenario uses its signature color (BEAR_COLOR, BASE_COLOR, BULL_COLOR).
    """
    # Collect data from all three scenario sheets
    scenario_names = ["bear", "base", "bull"]
    scenario_colors = [BEAR_COLOR, BASE_COLOR, BULL_COLOR]
    scenario_labels = ["Bear Case", "Base Case", "Bull Case"]

    # Metrics to extract from each scenario sheet
    metric_keys = [
        ("Total Revenue", "总收入", "Revenue ($M)"),
        ("EBITDA", "息税折旧摊销前利润", "EBITDA ($M)"),
        ("Equity Value", "股权价值", "Equity Value ($M)"),
        ("Per Share Value", "每股价值", "Price/Share ($)"),
    ]

    # Find the target sheet (we'll write helper data to the cover sheet or a new location)
    # Use the first available scenario sheet as the anchor, or the cover sheet
    # Better: write to the sheet specified by sheet_name
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    # Helper data area: far right columns
    helper_col_start = 30  # Column AD
    helper_row_start = 1

    # Write header row
    ws.cell(row=helper_row_start, column=helper_col_start, value="Scenario")
    for si, slabel in enumerate(scenario_labels):
        ws.cell(row=helper_row_start, column=helper_col_start + 1 + si, value=slabel)

    # For each metric, find values from each scenario sheet
    metric_values = {mk: [None, None, None] for mk, _, _ in metric_keys}

    for si, sc_name in enumerate(scenario_names):
        sc_sheet = _scenario_sheet_name(sc_name, "en")
        if sc_sheet not in wb.sheetnames:
            continue
        sc_ws = wb[sc_sheet]

        for row in range(1, sc_ws.max_row + 1):
            cell_val = sc_ws.cell(row=row, column=2).value
            if cell_val and isinstance(cell_val, str):
                cv = cell_val.strip()
                for mk_zh, mk_en, _ in metric_keys:
                    if mk_zh in cv or mk_en in cv:
                        # Get the last forecast year value (last data column)
                        val = sc_ws.cell(row=row, column=sc_ws.max_column).value
                        if val is not None:
                            # Use the key for storage
                            storage_key = mk_zh
                            metric_values[storage_key][si] = val
                        break

    # Write helper data for chart
    data_rows = []
    for mi, (mk_zh, mk_en, display_label) in enumerate(metric_keys):
        r = helper_row_start + 1 + mi
        ws.cell(row=r, column=helper_col_start, value=display_label)
        vals = metric_values[mk_zh]
        for si, v in enumerate(vals):
            if v is not None:
                ws.cell(row=r, column=helper_col_start + 1 + si, value=v)
        data_rows.append(r)

    n_metrics = len(data_rows)
    if n_metrics == 0:
        return

    # Categories (metric names)
    cats = Reference(ws, min_col=helper_col_start,
                     min_row=helper_row_start + 1,
                     max_row=helper_row_start + n_metrics)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Scenario Comparison: Bear / Base / Bull"
    chart.style = 2

    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.legend.position = 'b'
    chart.legend.includeInLayout = False

    chart.title = "Scenario Comparison: Bear / Base / Bull"
    chart.title.txPr = _chart_title_rich_text("Scenario Comparison: Bear / Base / Bull")

    chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    chart.gapWidth = CHART_GAP_WIDTH

    # Add one series per scenario (columns are scenario-oriented)
    for si in range(3):
        data_ref = Reference(ws, min_col=helper_col_start + 1 + si,
                             min_row=helper_row_start,
                             max_row=helper_row_start + n_metrics)
        chart.add_data(data_ref, titles_from_data=True)
        s = chart.series[si]
        s.graphicalProperties.solidFill = scenario_colors[si]

    chart.set_categories(cats)

    # Y-axis formatting
    chart.y_axis.numFmt = '#,##0'
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    # Data labels
    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = '#,##0'
        s.dLbls.txPr = RichText(
            p=[ChartParagraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
                ),
            )]
        )

    # Place chart below the stacked revenue chart
    ws.add_chart(chart, f"{get_column_letter(helper_col_start)}{helper_row_start + n_metrics + 2}")


def _add_dual_axis_combo_chart(wb, sheet_name, theme="classic"):
    """Add dual-axis combo chart: bar (Revenue $M) + line (EBITDA Margin %).

    Common in MS reports for showing absolute values alongside margins.
    Bar chart uses primary Y-axis (left), line chart uses secondary Y-axis (right).
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]

    year_header_row = 2
    col_start = 3

    # Find Total Revenue row and EBITDA row
    total_rev_row = None
    ebitda_row = None

    def _row_has_numeric_data(ws, row, col_start, col_end):
        for c in range(col_start, col_end + 1):
            val = ws.cell(row=row, column=c).value
            if val is not None:
                return True
        return False

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val and isinstance(cell_val, str):
            if "Total Revenue" in cell_val or "总收入" in cell_val:
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    total_rev_row = row
            elif cell_val.strip() == "EBITDA" or cell_val.strip() == "息税折旧摊销前利润":
                if _row_has_numeric_data(ws, row, col_start, ws.max_column):
                    ebitda_row = row

    if total_rev_row is None or ebitda_row is None:
        return

    # Find last data column
    last_col = col_start
    for c in range(col_start, ws.max_column + 1):
        if ws.cell(row=year_header_row, column=c).value is not None:
            last_col = c

    # Categories (year headers)
    cats = Reference(ws, min_col=col_start, max_col=last_col,
                     min_row=year_header_row, max_row=year_header_row)

    # Create bar chart for Revenue (primary axis)
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.grouping = "clustered"
    bar_chart.title = "Revenue ($M) vs EBITDA Margin (%)"
    bar_chart.style = 2
    bar_chart.y_axis.title = "$M"
    bar_chart.y_axis.numFmt = '#,##0'
    bar_chart.y_axis.axId = 100

    bar_chart.width = CHART_WIDTH_CM
    bar_chart.height = CHART_HEIGHT_CM
    bar_chart.legend.position = 'b'
    bar_chart.legend.includeInLayout = False

    bar_chart.title = "Revenue ($M) vs EBITDA Margin (%)"
    bar_chart.title.txPr = _chart_title_rich_text("Revenue ($M) vs EBITDA Margin (%)")

    bar_chart.legend.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LEGEND_SIZE)
            ),
        )]
    )

    bar_chart.gapWidth = CHART_GAP_WIDTH

    # Revenue series (bar)
    rev_data = Reference(ws, min_col=col_start, max_col=last_col,
                         min_row=total_rev_row, max_row=total_rev_row)
    bar_chart.add_data(rev_data, titles_from_data=False, from_rows=True)
    bar_chart.set_categories(cats)

    s_rev = bar_chart.series[0]
    s_rev.title = SeriesLabel(v="Revenue ($M)")
    s_rev.graphicalProperties.solidFill = CHART_PALETTE[0]

    # Data labels for revenue bars
    s_rev.dLbls = DataLabelList()
    s_rev.dLbls.showVal = True
    s_rev.dLbls.numFmt = '#,##0'
    s_rev.dLbls.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
            ),
        )]
    )

    bar_chart.y_axis.delete = False
    bar_chart.x_axis.delete = False

    # Create line chart for EBITDA Margin (secondary axis)
    # Write EBITDA Margin helper data
    helper_col_start = 27  # Column AA
    helper_row = max(ws.max_row + 2, 80)

    ws.cell(row=helper_row, column=helper_col_start, value="EBITDA Margin (helper)")
    for ci in range(col_start, last_col + 1):
        cl = get_column_letter(ci)
        hc = get_column_letter(helper_col_start + ci - col_start + 1)
        formula = f"=IF({cl}{total_rev_row}=0,0,{cl}{ebitda_row}/{cl}{total_rev_row})"
        c = ws.cell(row=helper_row, column=helper_col_start + ci - col_start + 1,
                    value=formula)
        c.number_format = '0.0%'

    # Copy year headers to helper area
    for ci in range(last_col - col_start + 1):
        src_val = ws.cell(row=year_header_row, column=col_start + ci).value
        ws.cell(row=helper_row - 1, column=helper_col_start + 1 + ci, value=src_val)

    helper_last_col = helper_col_start + (last_col - col_start + 1)

    line_chart = LineChart()
    line_chart.y_axis.title = "EBITDA Margin (%)"
    line_chart.y_axis.numFmt = '0%'
    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"

    margin_data = Reference(ws, min_col=helper_col_start + 1, max_col=helper_last_col,
                           min_row=helper_row, max_row=helper_row)
    line_chart.add_data(margin_data, titles_from_data=False, from_rows=True)

    # Use same categories as bar chart
    cats_helper = Reference(ws, min_col=helper_col_start + 1, max_col=helper_last_col,
                            min_row=helper_row - 1, max_row=helper_row - 1)
    line_chart.set_categories(cats_helper)

    s_margin = line_chart.series[0]
    s_margin.title = SeriesLabel(v="EBITDA Margin (%)")
    s_margin.graphicalProperties.line.solidFill = CHART_PALETTE[3]
    s_margin.graphicalProperties.line.width = CHART_LINE_WIDTH
    s_margin.marker.symbol = "circle"
    s_margin.marker.size = 5
    s_margin.marker.graphicalProperties.solidFill = CHART_PALETTE[3]

    # Data labels for margin line
    s_margin.dLbls = DataLabelList()
    s_margin.dLbls.showVal = True
    s_margin.dLbls.numFmt = '0.0%'
    s_margin.dLbls.txPr = RichText(
        p=[ChartParagraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=CHART_LABEL_SIZE, b=True)
            ),
        )]
    )

    line_chart.y_axis.delete = False
    line_chart.x_axis.delete = False

    # Overlay line chart on bar chart
    bar_chart += line_chart

    # Place chart below the stacked revenue chart
    anchor_col = last_col + 2
    ws.add_chart(bar_chart, f"{get_column_letter(anchor_col)}52")


# ============================================================
# 13. MAIN ENTRY POINT
# ============================================================

def make_financial_model(data: dict, output_path: str,
                         theme: str = "classic",
                         language: str = "zh") -> str:
    """
    Generate a Morgan Stanley style DCF financial model.

    Args:
        data: Model data dictionary (see sample_data() for structure)
        output_path: Path to save the .xlsx file
        theme: Color theme (reserved for future use)
        language: "zh" for Chinese, "en" for English

    Returns:
        Absolute path to the generated file
    """
    wb = Workbook()

    # 1. Cover sheet
    write_cover_sheet(wb, data, lang=language)

    # 2. Key Operational KPIs (optional, triggered by data["operational_kpis"])
    write_kpi_sheet(wb, data, lang=language)

    # 3. WACC Decomposition (must come before DCF for cross-sheet references)
    wacc_sheet_name = t("wacc_decomp", language)
    wacc_ws, wacc_result_row = write_wacc_sheet(wb, data, {}, lang=language)

    # 4. Three DCF sheets (with WACC cross-sheet reference)
    dcf_layouts = {}
    for scenario in ["bear", "base", "bull"]:
        layout = write_dcf_sheet(wb, data, scenario, lang=language,
                                  wacc_sheet_name=wacc_sheet_name,
                                  wacc_result_row=wacc_result_row)
        dcf_layouts[scenario] = layout

    # 5. Sensitivity Analysis
    write_sensitivity_sheet(wb, data, dcf_layouts, lang=language)

    # 6. SOTP Valuation (optional, triggered by data["sotp"])
    write_sotp_sheet(wb, data, lang=language)

    # 7. Comparable Companies (enhanced with multi-market grouping)
    write_comps_sheet(wb, data, lang=language)

    # 8. Historical PE Band (optional, triggered by data["pe_band"])
    write_pe_band_sheet(wb, data, lang=language)

    # 9. Second pass: fill cover sheet formulas
    fill_cover_formulas(wb, data, dcf_layouts, lang=language)

    # 10. Add charts to relevant sheets
    # Revenue & EBITDA bar chart + Margin line chart on Base DCF sheet
    base_sheet_name = _scenario_sheet_name("base", language)
    _add_revenue_ebitda_chart(wb, base_sheet_name, theme)
    _add_margin_line_chart(wb, base_sheet_name, theme)

    # Stacked revenue breakdown + dual-axis combo on Base DCF sheet
    _add_stacked_revenue_chart(wb, base_sheet_name, theme)
    _add_dual_axis_combo_chart(wb, base_sheet_name, theme)

    # SOTP pie chart + donut chart
    if data.get("sotp"):
        _add_sotp_pie_chart(wb, "SOTP", theme)
        _add_donut_chart(wb, "SOTP", theme)

    # PE Band area chart
    if data.get("pe_band"):
        _add_pe_band_chart(wb, "PE Band", theme)

    # Grouped scenario comparison chart (on Base DCF sheet, reads all 3 scenarios)
    _add_grouped_scenario_chart(wb, base_sheet_name, theme)

    # Print settings
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    return output_path


# ============================================================
# 13. SAMPLE DATA (V7 with SOTP, KPIs, PE Band, multi-market comps)
# ============================================================

def sample_data() -> dict:
    """Return complete sample data for Musk Empire Holdings DCF model.

    V7: Adds SOTP valuation, operational KPIs, PE Band, multi-market comps.
    """
    return {
        "company_name": "Musk Empire Holdings",
        "ticker": "MUSK",
        "report_date": "2026-06-12",
        "analyst": "Aurora Capital Research",
        "currency": "USD",
        "shares_outstanding": 1000,

        "business_overview": [
            {
                "name": "Autonomy & AI",
                "description": "Full self-driving software licensing and AI inference platform",
                "revenue_estimate": 8500,
            },
            {
                "name": "EV Manufacturing",
                "description": "Electric vehicle design, production and direct sales",
                "revenue_estimate": 12000,
            },
            {
                "name": "Energy & Storage",
                "description": "Solar panels, Powerwall, Megapack and grid-scale storage",
                "revenue_estimate": 6500,
            },
            {
                "name": "Financial Services",
                "description": "Insurance, lending and payment processing ecosystem",
                "revenue_estimate": 5000,
            },
            {
                "name": "Satellite Internet",
                "description": "Starlink LEO broadband constellation for consumer and enterprise",
                "revenue_estimate": 4200,
            },
            {
                "name": "Social Media",
                "description": "X platform advertising, subscription and data services",
                "revenue_estimate": 3000,
            },
            {
                "name": "Robotics",
                "description": "Humanoid robot (Optimus) for industrial and consumer use",
                "revenue_estimate": 1500,
            },
            {
                "name": "Neural Interface",
                "description": "Brain-computer interface technology for medical and consumer",
                "revenue_estimate": 800,
            },
            {
                "name": "Underground Transport",
                "description": "Boring Company tunnel infrastructure and loop transit systems",
                "revenue_estimate": 500,
            },
        ],

        "historical": {
            "years": ["FY21A", "FY22A", "FY23A", "FY24A", "FY25A"],
            "revenue": [5300, 7800, 11200, 15600, 20000],
            "ebitda": [530, 1014, 1680, 2500, 3600],
            "segments": {
                "Autonomy & AI": {"revenue": [800, 1400, 2200, 3200, 4500]},
                "EV Manufacturing": {"revenue": [2500, 3200, 4000, 5000, 5800]},
                "Energy & Storage": {"revenue": [800, 1000, 1400, 1800, 2200]},
                "Financial Services": {"revenue": [400, 500, 600, 800, 1000]},
                "Satellite Internet": {"revenue": [200, 400, 800, 1400, 2000]},
                "Social Media": {"revenue": [300, 500, 800, 1200, 1500]},
                "Robotics": {"revenue": [50, 100, 200, 400, 700]},
                "Neural Interface": {"revenue": [30, 50, 100, 200, 400]},
                "Underground Transport": {"revenue": [20, 50, 100, 200, 300]},
            },
        },

        "forecast_years": ["FY26E", "FY27E", "FY28E", "FY29E", "FY30E"],

        "scenarios": {
            "bear": {
                "label_zh": "熊市情景",
                "label_en": "Bear Case",
                "core_assumptions": "Conservative adoption, regulatory headwinds, margin compression",
                "segments": {
                    "Autonomy & AI": {
                        "fy25_base": 4500,
                        "growth_rates": [0.08, 0.06, 0.04, 0.03, 0.02],
                    },
                    "EV Manufacturing": {
                        "fy25_base": 5800,
                        "growth_rates": [0.05, 0.04, 0.03, 0.02, 0.01],
                    },
                    "Energy & Storage": {
                        "fy25_base": 2200,
                        "growth_rates": [0.06, 0.05, 0.04, 0.03, 0.02],
                    },
                    "Financial Services": {
                        "fy25_base": 1000,
                        "growth_rates": [0.05, 0.04, 0.03, 0.02, 0.02],
                    },
                    "Satellite Internet": {
                        "fy25_base": 2000,
                        "growth_rates": [0.12, 0.10, 0.08, 0.06, 0.05],
                    },
                    "Social Media": {
                        "fy25_base": 1500,
                        "growth_rates": [0.03, 0.02, 0.02, 0.01, 0.01],
                    },
                    "Robotics": {
                        "fy25_base": 700,
                        "growth_rates": [0.10, 0.08, 0.06, 0.05, 0.04],
                    },
                    "Neural Interface": {
                        "fy25_base": 400,
                        "growth_rates": [0.15, 0.12, 0.10, 0.08, 0.06],
                    },
                    "Underground Transport": {
                        "fy25_base": 300,
                        "growth_rates": [0.05, 0.04, 0.03, 0.03, 0.02],
                    },
                },
                "ebitda_margin": [0.18, 0.19, 0.19, 0.20, 0.20],
                "da_pct_revenue": 0.045,
                "capex_pct_revenue": 0.07,
                "nwc_pct_rev_change": 0.12,
                "tax_rate": 0.21,
                "wacc": {
                    "rf": 0.0440,
                    "erp": 0.0550,
                    "beta": 1.80,
                    "size_premium": 0.0050,
                    "country_risk": 0.0000,
                    "kd": 0.0650,
                    "tax_rate": 0.21,
                    "e_weight": 0.90,
                    "d_weight": 0.10,
                },
                "terminal_growth_rate": 0.020,
                "exit_multiple_ebitda": 8.0,
                "net_debt": 5000,
            },
            "base": {
                "label_zh": "基准情景",
                "label_en": "Base Case",
                "core_assumptions": "Steady growth across segments, margin expansion, FSD regulatory approval",
                "segments": {
                    "Autonomy & AI": {
                        "fy25_base": 4500,
                        "growth_rates": [0.15, 0.12, 0.10, 0.08, 0.06],
                    },
                    "EV Manufacturing": {
                        "fy25_base": 5800,
                        "growth_rates": [0.10, 0.08, 0.07, 0.05, 0.04],
                    },
                    "Energy & Storage": {
                        "fy25_base": 2200,
                        "growth_rates": [0.12, 0.10, 0.08, 0.07, 0.05],
                    },
                    "Financial Services": {
                        "fy25_base": 1000,
                        "growth_rates": [0.10, 0.08, 0.07, 0.06, 0.05],
                    },
                    "Satellite Internet": {
                        "fy25_base": 2000,
                        "growth_rates": [0.20, 0.18, 0.15, 0.12, 0.10],
                    },
                    "Social Media": {
                        "fy25_base": 1500,
                        "growth_rates": [0.08, 0.07, 0.06, 0.05, 0.04],
                    },
                    "Robotics": {
                        "fy25_base": 700,
                        "growth_rates": [0.25, 0.20, 0.15, 0.12, 0.10],
                    },
                    "Neural Interface": {
                        "fy25_base": 400,
                        "growth_rates": [0.30, 0.25, 0.20, 0.15, 0.12],
                    },
                    "Underground Transport": {
                        "fy25_base": 300,
                        "growth_rates": [0.10, 0.08, 0.07, 0.06, 0.05],
                    },
                },
                "ebitda_margin": [0.22, 0.24, 0.25, 0.26, 0.27],
                "da_pct_revenue": 0.035,
                "capex_pct_revenue": 0.055,
                "nwc_pct_rev_change": 0.08,
                "tax_rate": 0.21,
                "wacc": {
                    "rf": 0.0440,
                    "erp": 0.0550,
                    "beta": 1.50,
                    "size_premium": 0.0030,
                    "country_risk": 0.0000,
                    "kd": 0.0550,
                    "tax_rate": 0.21,
                    "e_weight": 0.92,
                    "d_weight": 0.08,
                },
                "terminal_growth_rate": 0.030,
                "exit_multiple_ebitda": 12.0,
                "net_debt": 3000,
            },
            "bull": {
                "label_zh": "牛市情景",
                "label_en": "Bull Case",
                "core_assumptions": "Accelerated FSD rollout, robotics mass adoption, satellite dominance",
                "segments": {
                    "Autonomy & AI": {
                        "fy25_base": 4500,
                        "growth_rates": [0.25, 0.20, 0.15, 0.12, 0.10],
                    },
                    "EV Manufacturing": {
                        "fy25_base": 5800,
                        "growth_rates": [0.15, 0.12, 0.10, 0.08, 0.06],
                    },
                    "Energy & Storage": {
                        "fy25_base": 2200,
                        "growth_rates": [0.18, 0.15, 0.12, 0.10, 0.08],
                    },
                    "Financial Services": {
                        "fy25_base": 1000,
                        "growth_rates": [0.15, 0.12, 0.10, 0.08, 0.07],
                    },
                    "Satellite Internet": {
                        "fy25_base": 2000,
                        "growth_rates": [0.30, 0.25, 0.20, 0.15, 0.12],
                    },
                    "Social Media": {
                        "fy25_base": 1500,
                        "growth_rates": [0.12, 0.10, 0.08, 0.07, 0.06],
                    },
                    "Robotics": {
                        "fy25_base": 700,
                        "growth_rates": [0.40, 0.30, 0.25, 0.20, 0.15],
                    },
                    "Neural Interface": {
                        "fy25_base": 400,
                        "growth_rates": [0.50, 0.40, 0.30, 0.25, 0.20],
                    },
                    "Underground Transport": {
                        "fy25_base": 300,
                        "growth_rates": [0.15, 0.12, 0.10, 0.08, 0.07],
                    },
                },
                "ebitda_margin": [0.25, 0.28, 0.30, 0.31, 0.32],
                "da_pct_revenue": 0.030,
                "capex_pct_revenue": 0.050,
                "nwc_pct_rev_change": 0.06,
                "tax_rate": 0.21,
                "wacc": {
                    "rf": 0.0440,
                    "erp": 0.0550,
                    "beta": 1.20,
                    "size_premium": 0.0020,
                    "country_risk": 0.0000,
                    "kd": 0.0450,
                    "tax_rate": 0.21,
                    "e_weight": 0.95,
                    "d_weight": 0.05,
                },
                "terminal_growth_rate": 0.035,
                "exit_multiple_ebitda": 16.0,
                "net_debt": 1000,
            },
        },

        "comparable_companies": [
            {"name": "Tesla", "ticker": "TSLA", "market": "US", "mcap": 800000, "ev": 750000,
             "ev_rev_ltm": 5.2, "ev_rev_ntm": 4.8, "ev_ebitda_ltm": 25.0,
             "ev_ebitda_ntm": 20.0, "cagr_3y": 0.35, "note": "Closest comp"},
            {"name": "Apple", "ticker": "AAPL", "market": "US", "mcap": 3500000, "ev": 3300000,
             "ev_rev_ltm": 8.0, "ev_rev_ntm": 7.5, "ev_ebitda_ltm": 22.0,
             "ev_ebitda_ntm": 20.0, "cagr_3y": 0.08, "note": "Tech ecosystem"},
            {"name": "Rivian", "ticker": "RIVN", "market": "US", "mcap": 120000, "ev": 140000,
             "ev_rev_ltm": 8.5, "ev_rev_ntm": 5.2, "ev_ebitda_ltm": -25.0,
             "ev_ebitda_ntm": -15.0, "cagr_3y": 0.55, "note": "Pre-profit"},
            {"name": "BYD", "ticker": "1211.HK", "market": "HK", "mcap": 950000, "ev": 880000,
             "ev_rev_ltm": 1.8, "ev_rev_ntm": 1.5, "ev_ebitda_ltm": 12.0,
             "ev_ebitda_ntm": 10.0, "cagr_3y": 0.42, "note": "China EV leader"},
            {"name": "Tencent", "ticker": "0700.HK", "market": "HK", "mcap": 450000, "ev": 420000,
             "ev_rev_ltm": 4.5, "ev_rev_ntm": 4.0, "ev_ebitda_ltm": 18.0,
             "ev_ebitda_ntm": 16.0, "cagr_3y": 0.10, "note": "Internet ecosystem"},
            {"name": "NIO", "ticker": "NIO", "market": "HK", "mcap": 150000, "ev": 170000,
             "ev_rev_ltm": 2.5, "ev_rev_ntm": 2.0, "ev_ebitda_ltm": -10.0,
             "ev_ebitda_ntm": -5.0, "cagr_3y": 0.30, "note": "Premium EV"},
            {"name": "BYD", "ticker": "002594.SZ", "market": "A", "mcap": 950000, "ev": 880000,
             "ev_rev_ltm": 1.8, "ev_rev_ntm": 1.5, "ev_ebitda_ltm": 12.0,
             "ev_ebitda_ntm": 10.0, "cagr_3y": 0.42, "note": "A-share listing"},
            {"name": "CATL", "ticker": "300750.SZ", "market": "A", "mcap": 1100000, "ev": 1000000,
             "ev_rev_ltm": 2.2, "ev_rev_ntm": 1.8, "ev_ebitda_ltm": 15.0,
             "ev_ebitda_ntm": 12.0, "cagr_3y": 0.38, "note": "Battery leader"},
            {"name": "XPeng", "ticker": "XPEV", "market": "US", "mcap": 80000, "ev": 95000,
             "ev_rev_ltm": 3.0, "ev_rev_ntm": 2.2, "ev_ebitda_ltm": -12.0,
             "ev_ebitda_ntm": -6.0, "cagr_3y": 0.28, "note": "Tech-focused"},
        ],

        "cover_subtitle": "Musk Empire Holdings - DCF Valuation Model",
        "key_takeaways": [
            "Base case implies $XX per share, 15% upside to current price",
            "Autonomy & AI segment is key value driver with highest growth",
            "WACC sensitivity suggests +/- 20% valuation range",
        ],
        "investment_thesis": "Market leader in autonomous driving and clean energy ecosystem.",
        "risk_factors": [
            "Regulatory risk in autonomous driving approval",
            "Competition from traditional OEMs entering EV space",
            "Supply chain concentration risk",
        ],

        # V7: SOTP (Sum-of-the-Parts) Valuation
        "sotp": {
            "segments": [
                {
                    "name": "Tesla Auto",
                    "revenue_ltm": 77000,
                    "ebitda_ltm": 15400,
                    "ev_multiple": 8.0,
                    "ev": 123200,
                    "method": "EV/EBITDA",
                    "note": "Comparable median",
                },
                {
                    "name": "SpaceX Launch",
                    "revenue_ltm": 10000,
                    "ebitda_ltm": 2000,
                    "ev_multiple": 15.0,
                    "ev": 30000,
                    "method": "EV/Revenue",
                    "note": "Industry avg",
                },
                {
                    "name": "Starlink",
                    "revenue_ltm": 4200,
                    "ebitda_ltm": 840,
                    "ev_multiple": 12.0,
                    "ev": 50400,
                    "method": "EV/Revenue",
                    "note": "Satellite comp median",
                },
                {
                    "name": "Energy & Storage",
                    "revenue_ltm": 6500,
                    "ebitda_ltm": 1300,
                    "ev_multiple": 6.0,
                    "ev": 39000,
                    "method": "EV/EBITDA",
                    "note": "Clean energy median",
                },
                {
                    "name": "X (Social Media)",
                    "revenue_ltm": 3000,
                    "ebitda_ltm": 600,
                    "ev_multiple": 5.0,
                    "ev": 15000,
                    "method": "EV/Revenue",
                    "note": "Social media avg",
                },
                {
                    "name": "Financial Services",
                    "revenue_ltm": 5000,
                    "ebitda_ltm": 1250,
                    "ev_multiple": 4.0,
                    "ev": 20000,
                    "method": "EV/EBITDA",
                    "note": "Fintech median",
                },
                {
                    "name": "Robotics (Optimus)",
                    "revenue_ltm": 1500,
                    "ebitda_ltm": 225,
                    "ev_multiple": 20.0,
                    "ev": 30000,
                    "method": "EV/Revenue",
                    "note": "Growth premium",
                },
                {
                    "name": "Neural Interface",
                    "revenue_ltm": 800,
                    "ebitda_ltm": 80,
                    "ev_multiple": 25.0,
                    "ev": 20000,
                    "method": "EV/Revenue",
                    "note": "Biotech premium",
                },
                {
                    "name": "Underground Transport",
                    "revenue_ltm": 500,
                    "ebitda_ltm": 100,
                    "ev_multiple": 8.0,
                    "ev": 4000,
                    "method": "EV/EBITDA",
                    "note": "Infra median",
                },
            ],
            "net_debt": -28000,
            "minority_interest": 0,
            "shares_outstanding": 3200,
        },

        # V7: Operational KPIs
        "operational_kpis": {
            "metrics": [
                {
                    "name": "Tesla Deliveries",
                    "unit": "K units",
                    "historical": [1310, 1390, 1810, 1636, 1790],
                    "forecast": [2000, 2200, 2400, 2600, 2800],
                },
                {
                    "name": "Starlink Users",
                    "unit": "M",
                    "historical": [2.5, 4.0, 6.0, 10.3, 15.0],
                    "forecast": [20.0, 25.0, 30.0, 35.0, 40.0],
                },
                {
                    "name": "SpaceX Launches",
                    "unit": "Times",
                    "historical": [31, 96, 134, 180, 220],
                    "forecast": [260, 300, 340, 380, 420],
                },
                {
                    "name": "FSD Subscribers",
                    "unit": "M",
                    "historical": [0.5, 1.2, 2.5, 5.0, 8.0],
                    "forecast": [12.0, 16.0, 20.0, 24.0, 28.0],
                },
                {
                    "name": "Optimus Units",
                    "unit": "K units",
                    "historical": [0.1, 0.5, 2.0, 5.0, 10.0],
                    "forecast": [25.0, 50.0, 80.0, 120.0, 150.0],
                },
            ],
        },

        # V7: Historical PE Band
        "pe_band": {
            "current_pe": 35.0,
            "forward_pe": 28.0,
            "median_pe": 55.0,
            "pe_history": [
                {"period": "FY2021", "high": 120, "low": 45, "close": 95, "eps": 2.5},
                {"period": "FY2022", "high": 100, "low": 30, "close": 55, "eps": 3.2},
                {"period": "FY2023", "high": 80, "low": 25, "close": 48, "eps": 4.1},
                {"period": "FY2024", "high": 70, "low": 20, "close": 42, "eps": 5.0},
                {"period": "FY2025", "high": 65, "low": 22, "close": 38, "eps": 5.8},
            ],
        },
    }


# ============================================================
# 14. CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate Morgan Stanley style DCF financial model (XLSX) - V7 SOTP Enhanced"
    )
    parser.add_argument("output", help="Output .xlsx file path")
    parser.add_argument("--lang", choices=["zh", "en"], default="zh",
                        help="Language: zh (Chinese) or en (English)")
    parser.add_argument("--theme", choices=["classic"], default="classic",
                        help="Color theme (default: classic)")
    args = parser.parse_args()

    data = sample_data()
    output = make_financial_model(data, args.output, theme=args.theme, language=args.lang)
    print(f"Financial model generated: {output}")


if __name__ == "__main__":
    main()
